import webbrowser
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
from typing import List, Dict, Any, Optional

pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)

# ---------- 常量 ----------
FONT_SIZE_HTML = 12  # 可全局调整，也可通过函数参数覆盖


# ---------- 分类映射 ----------
def classify_main_type(val: str) -> str:
    """将产品主推类型映射为三大分类"""
    if val in ['TOP', 'TOP+']:
        return '标准化'
    elif val == 'NON-TOP':
        return '配置化'
    elif val == '定制':
        return '定制化'
    else:
        return '其他'


# ---------- 按销售大区统计 ----------
def compute_region_stats(df: pd.DataFrame) -> tuple:
    """
    按销售大区统计标准化/配置化/定制化面积总和，计算排序、占比、排名。
    返回 (sorted_regions, region_data_dict)
    """
    df_temp = df.copy()
    df_temp['主推分类'] = df_temp['产品主推类型（产品维度）'].apply(classify_main_type)

    region_class = df_temp.groupby(['销售大区', '主推分类'])['总面积'].sum().unstack(fill_value=0)
    for cat in ['标准化', '配置化', '定制化']:
        if cat not in region_class.columns:
            region_class[cat] = 0
    region_class['面积汇总'] = region_class.sum(axis=1)

    total_area = region_class['面积汇总'].sum()
    region_class['占比'] = (region_class['面积汇总'] / total_area * 100) if total_area else 0
    region_class_sorted = region_class.sort_values('面积汇总', ascending=False)
    region_class_sorted['排名'] = range(1, len(region_class_sorted) + 1)

    sorted_regions = region_class_sorted.index.tolist()
    region_data = {}
    for region in sorted_regions:
        row = region_class_sorted.loc[region]
        region_data[region] = {
            '面积汇总': row['面积汇总'],
            '标准化': row['标准化'],
            '配置化': row['配置化'],
            '定制化': row['定制化'],
            '占比': row['占比'],
            '排名': row['排名']
        }
    return sorted_regions, region_data


# ---------- 核心数据处理 ----------
def process_region_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    对已过滤好的DataFrame（仅含一个区域）进行处理：
    分组透视 → 添加小计（按平台类别）→ 插入顶部总计行 → 添加底部总计行
    新增『总计』列（最右侧），统计所有大区列之和
    平台按小计占比降序，平台内产品系列按系列汇总降序，系列内产品型号按面积降序
    """
    # 1. 按平台类别+产品系列+产品型号汇总面积和订单数
    grouped = df.groupby(['平台类别', '产品系列', '产品型号'], as_index=False).agg({
        '总面积': 'sum',
        '订单数量': 'sum'
    })

    # 2. 合并主推分类信息，并添加定制原因
    type_map = df[['产品型号', '产品主推类型（产品维度）', '主推分类']].drop_duplicates(subset='产品型号')
    grouped = grouped.merge(type_map, on='产品型号', how='left')
    grouped['定制原因'] = grouped['产品主推类型（产品维度）'].apply(lambda x: '定制' if x == '定制' else '')

    # 3. 透视成三列（标准化/配置化/定制化）
    pivot_df = grouped.pivot_table(
        index=['平台类别', '产品系列', '产品型号'],
        columns='主推分类',
        values='总面积',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    for cat in ['标准化', '配置化', '定制化']:
        if cat not in pivot_df.columns:
            pivot_df[cat] = 0

    pivot_df['面积汇总'] = pivot_df[['标准化', '配置化', '定制化']].sum(axis=1)

    # 合并定制原因
    custom_map = grouped[['平台类别', '产品系列', '产品型号', '定制原因']].drop_duplicates()
    pivot_df = pivot_df.merge(custom_map, on=['平台类别', '产品系列', '产品型号'], how='left')

    # --- 按销售大区统计明细 ---
    region_stats = compute_region_stats(df)
    sorted_regions, region_data = region_stats

    region_group = df.groupby(['平台类别', '产品系列', '产品型号', '销售大区', '主推分类']).agg({
        '总面积': 'sum'
    }).reset_index()

    region_pivot = region_group.pivot_table(
        index=['平台类别', '产品系列', '产品型号'],
        columns=['销售大区', '主推分类'],
        values='总面积',
        fill_value=0
    ).reset_index()

    # 展平列名
    key_cols = ['平台类别', '产品系列', '产品型号']
    keys = pd.DataFrame()
    for col in key_cols:
        found = False
        for c in region_pivot.columns:
            if col in str(c):
                keys[col] = region_pivot[c]
                found = True
                break
        if not found:
            raise KeyError(f"在 region_pivot 中找不到列: {col}")

    stat_cols = region_pivot.drop(columns=list(keys.columns))
    stat_cols.columns = [
        f"{col[0]}_{col[1]}" if isinstance(col, tuple) else str(col)
        for col in stat_cols.columns
    ]
    region_pivot_flat = pd.concat([keys, stat_cols], axis=1)

    # 合并大区数据
    pivot_df = pivot_df.merge(region_pivot_flat, on=key_cols, how='left')

    # 确保所有大区列都存在，并填充NaN为0
    region_cols = []
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            region_cols.append(f"{region}_{cat}")

    for col in region_cols:
        if col not in pivot_df.columns:
            pivot_df[col] = 0
        else:
            pivot_df[col] = pivot_df[col].fillna(0)

    # 新增『总计』列（计算所有大区列之和）
    pivot_df['总计'] = pivot_df[region_cols].sum(axis=1)

    # 调整列顺序
    base_cols = ['平台类别', '产品系列', '产品型号', '面积汇总', '标准化', '配置化', '定制化', '定制原因']
    final_cols = base_cols + region_cols + ['总计']
    for col in final_cols:
        if col not in pivot_df.columns:
            pivot_df[col] = 0

    final_df = pivot_df[final_cols]

    # ===== 取整 =====
    numeric_cols = ['面积汇总', '标准化', '配置化', '定制化'] + region_cols + ['总计']
    for col in numeric_cols:
        if col in final_df.columns:
            final_df[col] = final_df[col].round(0).astype(int)

    # 重新计算面积汇总（取整后三列之和）
    final_df['面积汇总'] = final_df['标准化'] + final_df['配置化'] + final_df['定制化']

    # ===== 新增排序逻辑 =====
    # 1. 计算每个平台的汇总面积（用于平台排序）
    platform_totals = final_df.groupby('平台类别')['面积汇总'].sum().to_dict()
    final_df['_platform_total'] = final_df['平台类别'].map(platform_totals)

    # 2. 计算每个平台内每个产品系列的汇总面积（用于系列排序）
    series_totals = final_df.groupby(['平台类别', '产品系列'])['面积汇总'].sum().to_dict()
    final_df['_series_total'] = final_df.apply(
        lambda r: series_totals.get((r['平台类别'], r['产品系列']), 0), axis=1
    )

    # 3. 按平台汇总降序 → 系列汇总降序 → 型号面积降序排序
    final_df = final_df.sort_values(
        ['_platform_total', '_series_total', '面积汇总'],
        ascending=[False, False, False]
    ).drop(columns=['_platform_total', '_series_total']).reset_index(drop=True)

    # ---- 计算总计值 ----
    total_area = final_df['面积汇总'].sum()
    total_std = final_df['标准化'].sum()
    total_cfg = final_df['配置化'].sum()
    total_cus = final_df['定制化'].sum()

    std_ratio = (total_std / total_area * 100) if total_area else 0
    cfg_ratio = (total_cfg / total_area * 100) if total_area else 0
    cus_ratio = (total_cus / total_area * 100) if total_area else 0
    col_std = f"标准化({std_ratio:.1f}%)"
    col_cfg = f"配置化({cfg_ratio:.1f}%)"
    col_cus = f"定制化({cus_ratio:.1f}%)"

    # ---- 构建最终行（包含明细、小计、顶部总计、底部总计） ----
    rows = []

    # ① 顶部总计行
    total_row_dict = {
        '平台类别': '总计',
        '产品系列': '',
        '产品型号': '',
        '面积汇总': total_area,
        '标准化': total_std,
        '配置化': total_cfg,
        '定制化': total_cus,
        '定制原因': ''
    }
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            col_name = f"{region}_{cat}"
            total_row_dict[col_name] = final_df[col_name].sum()
    total_row_dict['总计'] = final_df['总计'].sum()

    top_total_row = total_row_dict.copy()
    top_total_row['平台类别'] = ''
    top_total_row['产品系列'] = ''
    top_total_row['产品型号'] = ''
    rows.append(top_total_row)

    # ② 平台明细 + 小计（按排序后的顺序）
    # 此时 final_df 已按平台、系列、型号排序，直接 groupby 保留顺序
    for plat, group in final_df.groupby('平台类别', sort=False):
        for _, row in group.iterrows():
            rows.append(row.to_dict())
        # 小计行
        g_sum_area = group['面积汇总'].sum()
        g_sum_std = group['标准化'].sum()
        g_sum_cfg = group['配置化'].sum()
        g_sum_cus = group['定制化'].sum()
        ratio = (g_sum_area / total_area * 100) if total_area != 0 else 0
        ratio_text = f"占比 {ratio:.2f}%"
        subtotal = {
            '平台类别': f"{plat}汇总",
            '产品系列': ratio_text,
            '产品型号': '',
            '面积汇总': g_sum_area,
            '标准化': g_sum_std,
            '配置化': g_sum_cfg,
            '定制化': g_sum_cus,
            '定制原因': ''
        }
        for region in sorted_regions:
            for cat in ['标准化', '配置化', '定制化']:
                col_name = f"{region}_{cat}"
                if col_name in group.columns:
                    subtotal[col_name] = group[col_name].sum()
                else:
                    subtotal[col_name] = 0
        subtotal['总计'] = sum(subtotal[col] for col in region_cols)
        rows.append(subtotal)

    # ③ 底部总计行
    rows.append(total_row_dict)

    result_df = pd.DataFrame(rows)

    # 重命名列（含占比的标题）
    new_cols = ['平台类别', '产品系列', '产品型号', '面积汇总', col_std, col_cfg, col_cus, '定制原因']
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            new_cols.append(f"{region}_{cat}")
    new_cols.append('总计')
    result_df.columns = new_cols

    # 存储元数据
    result_df.attrs['sorted_regions'] = sorted_regions
    result_df.attrs['region_data'] = region_data
    result_df.attrs['region_area'] = {r: region_data[r]['面积汇总'] for r in sorted_regions}
    result_df.attrs['region_ratio'] = {r: region_data[r]['占比'] for r in sorted_regions}
    result_df.attrs['region_rank'] = {r: region_data[r]['排名'] for r in sorted_regions}

    return result_df

# ---------- Excel 合并与样式 ----------
def apply_merges(
    sheet,
    start_row: int = 3,
    sorted_regions: List[str] = None,
    region_data: Dict[str, Any] = None,
    col_names: List[str] = None
) -> None:
    """
    合并单元格、设置表头、添加边框和背景色。
    第1-2行：表头（前8列第一二行合并，大区列第一行合并三列，『总计』列第一二行合并）
    第3行：顶部总计行（合并前三列，浅蓝色背景）
    第4行起：数据行（明细、小计）
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row < start_row:
        return

    # ---------- 定义边框样式 ----------
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ---------- 写入表头 ----------
    if col_names is None:
        col_names = [sheet.cell(row=1, column=i).value for i in range(1, 9)]
        if not any(col_names):
            col_names = ['平台类别', '产品系列', '产品型号', '面积汇总', '标准化', '配置化', '定制化', '定制原因']

    # 第1-2行：前8列合并并写入标题
    for col_idx, name in enumerate(col_names[:8], start=1):
        # 合并第一二行
        sheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = name
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 大区表头（从第9列开始）
    if sorted_regions and region_data:
        start_col = 9
        # 记录『总计』列的位置（在大区列之后）
        total_col_index = start_col + len(sorted_regions) * 3

        for i, region in enumerate(sorted_regions):
            col_start = start_col + i * 3
            col_end = col_start + 2
            # 第1行：合并三列，显示大区信息
            sheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            info = region_data.get(region, {})
            rank = info.get('排名', 0)
            area = info.get('面积汇总', 0)
            ratio = info.get('占比', 0)
            display_text = f"TOP{rank:.0f} {region}\n{area:.2f}㎡ ({ratio:.1f}%)"
            cell = sheet.cell(row=1, column=col_start)
            cell.value = display_text
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            # 第2行：子标题
            for j, sub in enumerate(['标准化', '配置化', '定制化']):
                cell = sheet.cell(row=2, column=col_start + j)
                cell.value = sub
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # ---- 『总计』列表头（第一二行合并） ----
        sheet.merge_cells(start_row=1, start_column=total_col_index, end_row=2, end_column=total_col_index)
        cell = sheet.cell(row=1, column=total_col_index)
        cell.value = '总计'
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # ---------- 设置所有单元格居中，并添加边框 ----------
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # ---------- 设置顶部总计行（第3行）合并前三列，填"总计"，浅蓝色背景 ----------
    if max_row >= 3:
        # 合并前三列
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
        cell = sheet.cell(row=3, column=1)
        cell.value = '总计'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 其余列浅蓝色背景（已有数据）
        for col in range(4, max_col + 1):
            cell = sheet.cell(row=3, column=col)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)

    # ---------- 合并明细行（从第4行开始） ----------
    detail_rows = []
    for r in range(start_row, max_row + 1):
        plat_val = sheet.cell(row=r, column=1).value
        if plat_val is None:
            continue
        if isinstance(plat_val, str) and (plat_val.endswith('汇总') or plat_val == '总计'):
            continue
        # 跳过顶部总计行（第3行，已处理）
        if r == 3:
            continue
        series_val = sheet.cell(row=r, column=2).value
        model_val = sheet.cell(row=r, column=3).value
        detail_rows.append((r, plat_val, series_val, model_val))

    # ---------- 加粗小计/总计，合并小计行产品系列+产品型号 ----------
    for row_idx in range(start_row, max_row + 1):
        plat_cell = sheet.cell(row=row_idx, column=1)
        if plat_cell.value and isinstance(plat_cell.value, str):
            val = plat_cell.value
            if val.endswith('汇总'):
                # 小计行背景浅绿色，加粗
                sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
                b_val = sheet.cell(row=row_idx, column=2).value
                merged_cell = sheet.cell(row=row_idx, column=2)
                merged_cell.value = b_val
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif val == '总计':
                # 底部总计行：合并第一至三列，浅蓝色背景，加粗
                sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                merged_cell = sheet.cell(row=row_idx, column=1)
                merged_cell.value = '总计'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ---------- 平台类别合并（列1） ----------
    if detail_rows:
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))
        for plat, rows in platform_groups.items():
            start_r = rows[0][0]
            end_r = rows[-1][0]
            if len(rows) > 1:
                sheet.merge_cells(start_row=start_r, start_column=1,
                                  end_row=end_r, end_column=1)
                merged_cell = sheet.cell(row=start_r, column=1)
                merged_cell.value = plat

    # ---------- 产品系列（列2）和产品型号（列3）合并 ----------
    if detail_rows:
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))
        for plat, rows in platform_groups.items():
            idx = 0
            while idx < len(rows):
                row_idx, series_val, model_val = rows[idx]
                if series_val is None:
                    idx += 1
                    continue
                end_idx = idx
                while end_idx < len(rows) and str(rows[end_idx][1]) == str(series_val):
                    end_idx += 1
                if end_idx - idx > 1:
                    start_r = rows[idx][0]
                    end_r = rows[end_idx - 1][0]
                    sheet.merge_cells(start_row=start_r, start_column=2,
                                      end_row=end_r, end_column=2)
                    merged_cell = sheet.cell(row=start_r, column=2)
                    merged_cell.value = series_val
                sub_rows = rows[idx:end_idx]
                j = 0
                while j < len(sub_rows):
                    r_j, _, model_val_j = sub_rows[j]
                    if model_val_j is None:
                        j += 1
                        continue
                    j2 = j
                    while j2 < len(sub_rows) and sub_rows[j2][2] == model_val_j:
                        j2 += 1
                    if j2 - j > 1:
                        start_r_model = sub_rows[j][0]
                        end_r_model = sub_rows[j2 - 1][0]
                        sheet.merge_cells(start_row=start_r_model, start_column=3,
                                          end_row=end_r_model, end_column=3)
                        merged_cell = sheet.cell(row=start_r_model, column=3)
                        merged_cell.value = model_val_j
                    j = j2
                idx = end_idx

    # ---------- 设置列宽 ----------
    col_widths = {1: 22, 2: 18, 3: 12, 4: 14, 5: 16, 6: 16, 7: 16, 8: 14}
    for col, width in col_widths.items():
        if col <= max_col:
            sheet.column_dimensions[get_column_letter(col)].width = width
    for col in range(9, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14


# ---------- HTML 生成 ----------
def df_to_html_with_merges(df: pd.DataFrame, title: str, font_size: int = FONT_SIZE_HTML) -> str:
    """
    将含小计/总计的 DataFrame 转换为 HTML 表格，
    自动合并单元格，支持顶部总计行和底部总计行合并，
    并设置背景色（总计浅蓝，小计浅绿）。
    """
    rows = df.to_dict('records')
    if not rows:
        return "<p>无数据</p>"

    cols = df.columns.tolist()
    sorted_regions = df.attrs.get('sorted_regions', [])
    region_data = df.attrs.get('region_data', {})

    # 查找列
    std_col = next((c for c in cols if '标准化' in c and not any(c.startswith(r) for r in sorted_regions)), None)
    cfg_col = next((c for c in cols if '配置化' in c and not any(c.startswith(r) for r in sorted_regions)), None)
    cus_col = next((c for c in cols if '定制化' in c and not any(c.startswith(r) for r in sorted_regions)), None)
    area_col = '面积汇总'
    custom_col = '定制原因'
    total_col = '总计'   # 新增列

    # 构建两行表头（前8列使用rowspan=2）
    thead_parts = ['<thead><tr>']
    # 前8列：rowspan=2
    for col_name in ['平台类别', '产品系列', '产品型号', '面积汇总', std_col, cfg_col, cus_col, custom_col]:
        thead_parts.append(f'<th rowspan="2">{col_name}</th>')
    # 大区列（第一行合并三列）
    for region in sorted_regions:
        info = region_data.get(region, {'排名': 0, '面积汇总': 0, '占比': 0})
        rank = info['排名']
        area = info['面积汇总']
        ratio = info['占比']
        display_text = f"TOP{rank:.0f} - {region}<br>{area:.2f}㎡ - ({ratio:.1f}%)"
        thead_parts.append(f'<th colspan="3">{display_text}</th>')
    # 『总计』列：第一行跨两行
    thead_parts.append(f'<th rowspan="2">总计</th>')
    thead_parts.append('</tr><tr>')
    # 第二行：大区的子标题
    for region in sorted_regions:
        thead_parts.append('<th>标准化</th><th>配置化</th><th>定制化</th>')
    thead_parts.append('</tr></thead>')
    thead = ''.join(thead_parts)

    # 构建 tbody
    tbody_rows = []
    row_types = []
    for row in rows:
        plat = row['平台类别']
        if plat == '总计':
            row_types.append('total')
        elif isinstance(plat, str) and plat.endswith('汇总'):
            row_types.append('subtotal')
        elif plat == '':   # 顶部总计行（第一列为空）
            row_types.append('top_total')
        else:
            row_types.append('detail')

    i = 0
    total_rows = len(rows)

    while i < total_rows:
        row = rows[i]
        row_type = row_types[i]

        if row_type == 'top_total':
            # 顶部总计行：前三列合并显示"总计"，其余显示数值
            tr = '<tr class="top-total">'
            # 前三列合并
            tr += '<td colspan="3">总计</td>'
            # 面积汇总、标准化、配置化、定制化、定制原因
            area_val = row.get(area_col, 0)
            std_val = row.get(std_col, 0)
            cfg_val = row.get(cfg_col, 0)
            cus_val = row.get(cus_col, 0)
            custom_val = row.get(custom_col, '')
            tr += f'<td>{"" if area_val==0 else str(int(area_val))}</td>'
            tr += f'<td>{"" if std_val==0 else str(int(std_val))}</td>'
            tr += f'<td>{"" if cfg_val==0 else str(int(cfg_val))}</td>'
            tr += f'<td>{"" if cus_val==0 else str(int(cus_val))}</td>'
            tr += f'<td>{custom_val}</td>'
            # 大区列
            for region in sorted_regions:
                std_col_r = f"{region}_标准化"
                cfg_col_r = f"{region}_配置化"
                cus_col_r = f"{region}_定制化"
                std_val_r = row.get(std_col_r, 0)
                cfg_val_r = row.get(cfg_col_r, 0)
                cus_val_r = row.get(cus_col_r, 0)
                tr += f'<td>{"" if std_val_r==0 else str(int(std_val_r))}</td>'
                tr += f'<td>{"" if cfg_val_r==0 else str(int(cfg_val_r))}</td>'
                tr += f'<td>{"" if cus_val_r==0 else str(int(cus_val_r))}</td>'
            # 总计列
            total_val = row.get(total_col, 0)
            tr += f'<td>{"" if total_val==0 else str(int(total_val))}</td>'
            tr += '</tr>'
            tbody_rows.append(tr)
            i += 1
            continue

        if row_type in ('subtotal', 'total'):
            plat_val = row['平台类别']
            if row_type == 'subtotal':
                series_val = row['产品系列']
                css_class = 'subtotal'
            else:
                series_val = ''
                css_class = 'total'
            custom_val = ''
            area_val = row.get(area_col, 0)
            std_val = row.get(std_col, 0)
            cfg_val = row.get(cfg_col, 0)
            cus_val = row.get(cus_col, 0)
            total_val = row.get(total_col, 0)

            area_disp = '' if area_val == 0 else str(int(area_val))
            std_disp = '' if std_val == 0 else str(int(std_val))
            cfg_disp = '' if cfg_val == 0 else str(int(cfg_val))
            cus_disp = '' if cus_val == 0 else str(int(cus_val))
            total_disp = '' if total_val == 0 else str(int(total_val))

            tr = f'<tr class="{css_class}">'
            if row_type == 'total':
                tr += f'<td colspan="3">{plat_val}</td>'
            else:
                tr += f'<td>{plat_val}</td>'
                tr += f'<td colspan="2">{series_val}</td>'
            tr += f'<td>{area_disp}</td>'
            tr += f'<td>{std_disp}</td>'
            tr += f'<td>{cfg_disp}</td>'
            tr += f'<td>{cus_disp}</td>'
            tr += f'<td>{custom_val}</td>'
            for region in sorted_regions:
                std_col_r = f"{region}_标准化"
                cfg_col_r = f"{region}_配置化"
                cus_col_r = f"{region}_定制化"
                std_val_r = row.get(std_col_r, 0)
                cfg_val_r = row.get(cfg_col_r, 0)
                cus_val_r = row.get(cus_col_r, 0)
                tr += f'<td>{"" if std_val_r==0 else str(int(std_val_r))}</td>'
                tr += f'<td>{"" if cfg_val_r==0 else str(int(cfg_val_r))}</td>'
                tr += f'<td>{"" if cus_val_r==0 else str(int(cus_val_r))}</td>'
            tr += f'<td>{total_disp}</td>'
            tr += '</tr>'
            tbody_rows.append(tr)
            i += 1
            continue

        # 明细行
        curr_plat = row['平台类别']
        start_i = i
        while i < total_rows and row_types[i] == 'detail' and rows[i]['平台类别'] == curr_plat:
            i += 1
        end_i = i - 1
        plat_rowspan = end_i - start_i + 1

        j = start_i
        while j <= end_i:
            curr_series = rows[j]['产品系列']
            series_start = j
            while j <= end_i and rows[j]['产品系列'] == curr_series:
                j += 1
            series_end = j - 1
            series_rowspan = series_end - series_start + 1

            k = series_start
            while k <= series_end:
                curr_model = rows[k]['产品型号']
                model_start = k
                while k <= series_end and rows[k]['产品型号'] == curr_model:
                    k += 1
                model_end = k - 1
                model_rowspan = model_end - model_start + 1

                for idx in range(model_start, model_end + 1):
                    row_data = rows[idx]
                    if idx == model_start:
                        custom_val = row_data.get(custom_col, '')
                        plat_td = f'<td rowspan="{plat_rowspan}">{curr_plat}</td>' if idx == start_i else ''
                        series_td = f'<td rowspan="{series_rowspan}">{curr_series}</td>' if idx == series_start else ''
                        model_td = f'<td rowspan="{model_rowspan}">{curr_model}</td>' if idx == model_start else ''
                    else:
                        custom_val = ''
                        plat_td = ''
                        series_td = ''
                        model_td = ''

                    area_val = row_data.get(area_col, 0)
                    std_val = row_data.get(std_col, 0)
                    cfg_val = row_data.get(cfg_col, 0)
                    cus_val = row_data.get(cus_col, 0)
                    total_val = row_data.get(total_col, 0)

                    area_disp = '' if area_val == 0 else str(int(area_val))
                    std_disp = '' if std_val == 0 else str(int(std_val))
                    cfg_disp = '' if cfg_val == 0 else str(int(cfg_val))
                    cus_disp = '' if cus_val == 0 else str(int(cus_val))
                    total_disp = '' if total_val == 0 else str(int(total_val))

                    tr = '<tr>'
                    if plat_td:
                        tr += plat_td
                    if series_td:
                        tr += series_td
                    if model_td:
                        tr += model_td
                    tr += f'<td>{area_disp}</td>'
                    tr += f'<td>{std_disp}</td>'
                    tr += f'<td>{cfg_disp}</td>'
                    tr += f'<td>{cus_disp}</td>'
                    tr += f'<td>{custom_val}</td>'
                    for region in sorted_regions:
                        std_col_r = f"{region}_标准化"
                        cfg_col_r = f"{region}_配置化"
                        cus_col_r = f"{region}_定制化"
                        std_val_r = row_data.get(std_col_r, 0)
                        cfg_val_r = row_data.get(cfg_col_r, 0)
                        cus_val_r = row_data.get(cus_col_r, 0)
                        tr += f'<td>{"" if std_val_r==0 else str(int(std_val_r))}</td>'
                        tr += f'<td>{"" if cfg_val_r==0 else str(int(cfg_val_r))}</td>'
                        tr += f'<td>{"" if cus_val_r==0 else str(int(cus_val_r))}</td>'
                    tr += f'<td>{total_disp}</td>'
                    tr += '</tr>'
                    tbody_rows.append(tr)

                k = model_end + 1
            j = series_end + 1
        i = end_i + 1

    html = f"""
    <div class="table-container">
        <h2>{title}</h2>
        <table style="font-size: {font_size}px; border-collapse: collapse; width: 100%;">
            {thead}
            <tbody>
                {''.join(tbody_rows)}
            </tbody>
        </table>
    </div>
    """
    return html


# ---------- 主函数 ----------
def main() -> None:
    data_path = 'data/接单明细V1.1.xlsx'
    data = pd.read_excel(data_path)
    print("原始列：", data.columns.tolist())

    titles = ['创建日期', '区域', '销售大区', '省份/国家', '国家', '订单数量', '总面积',
              '订单推送金额-RMB', '签单金额（万元）', '产品系列', '产品型号', '产品间距',
              '产品主推类型（产品维度）', '修正后的产品线', '修正后的业务产品线', '品牌',
              '业务产品线', '旗舰类别', '平台类别', '定位']
    data = data[titles]

    data['主推分类'] = data['产品主推类型（产品维度）'].apply(classify_main_type)

    domestic = data[data['区域'] == '国内'].copy()
    international = data[data['区域'] == '国际'].copy()

    df_dom = process_region_data(domestic)
    df_int = process_region_data(international)

    # ---------- 写入Excel ----------
    output_file = '分组统计结果.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_dom_excel = df_dom.copy()
        df_int_excel = df_int.copy()
        numeric_cols_dom = [col for col in df_dom_excel.columns if col not in ['平台类别', '产品系列', '产品型号', '定制原因']]
        for col in numeric_cols_dom:
            df_dom_excel[col] = df_dom_excel[col].replace(0, None)
        numeric_cols_int = [col for col in df_int_excel.columns if col not in ['平台类别', '产品系列', '产品型号', '定制原因']]
        for col in numeric_cols_int:
            df_int_excel[col] = df_int_excel[col].replace(0, None)

        df_dom_excel.to_excel(writer, sheet_name='国内', index=False, header=False, startrow=2)
        df_int_excel.to_excel(writer, sheet_name='国际', index=False, header=False, startrow=2)
        workbook = writer.book

        sorted_regions_dom = df_dom.attrs.get('sorted_regions', [])
        region_data_dom = df_dom.attrs.get('region_data', {})
        sorted_regions_int = df_int.attrs.get('sorted_regions', [])
        region_data_int = df_int.attrs.get('region_data', {})

        for sheet_name in ['国内', '国际']:
            sheet = workbook[sheet_name]
            if sheet_name == '国内':
                sorted_regions = sorted_regions_dom
                region_data = region_data_dom
                col_names = df_dom.columns.tolist()
            else:
                sorted_regions = sorted_regions_int
                region_data = region_data_int
                col_names = df_int.columns.tolist()
            apply_merges(sheet, start_row=3, sorted_regions=sorted_regions,
                         region_data=region_data, col_names=col_names)

    print(f"Excel保存完成！文件：{output_file}")

    # ---------- 生成网页HTML ----------
    dom_html = df_to_html_with_merges(df_dom, "国内", font_size=FONT_SIZE_HTML)
    int_html = df_to_html_with_merges(df_int, "国际", font_size=FONT_SIZE_HTML)

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>分组统计结果</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h2 {{ text-align: center; color: #333; }}
            .table-container {{ margin-bottom: 40px; overflow-x: auto; }}
            table {{ border-collapse: collapse; width: 100%; margin: 0 auto; }}
            th, td {{ border: 1px solid #000; padding: 6px 12px; text-align: center; }}
            th {{ background-color: #4A90D9; color: white; font-weight: bold; }}
            .top-total td {{ background-color: #D9E1F2; font-weight: bold; }}
            .total td {{ background-color: #D9E1F2; font-weight: bold; }}
            .subtotal td {{ background-color: #E2EFDA; font-weight: bold; }}
        </style>
    </head>
    <body>
        {dom_html}
        {int_html}
    </body>
    </html>
    """

    html_file = '分组统计结果.html'
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    webbrowser.open(html_file)
    print(f"网页已生成并打开：{html_file}")


if __name__ == '__main__':
    main()