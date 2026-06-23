import webbrowser
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)


# ---------- 分类映射 ----------
def classify_main_type(val):
    if val in ['TOP', 'TOP+']:
        return '标准化'
    elif val == 'NON-TOP':
        return '配置化'
    elif val == '定制':
        return '定制化'
    else:
        return '其他'


# ---------- 处理单个区域数据（定制原因放在最右侧） ----------
def process_region_data(df):
    """
    对已过滤好的DataFrame（仅含一个区域）进行处理：
    分组透视 → 添加小计（按平台类别）→ 添加总计 → 返回含小计/总计的DataFrame
    定制原因列放在最右侧
    """
    # 1. 按平台类别+产品系列+产品型号汇总面积和订单数
    grouped = df.groupby(['平台类别', '产品系列', '产品型号'], as_index=False).agg({
        '总面积': 'sum',
        '订单数量': 'sum'
    })

    # 2. 合并主推分类信息（每个产品型号唯一），并添加定制原因
    type_map = df[['产品型号', '产品主推类型（产品维度）', '主推分类']].drop_duplicates(subset='产品型号')
    grouped = grouped.merge(type_map, on='产品型号', how='left')
    grouped['定制原因'] = grouped['产品主推类型（产品维度）'].apply(lambda x: '定制' if x == '定制' else '')

    # 3. 透视成三列（标准化/配置化/定制化）
    pivot_df = grouped.pivot_table(
        index=['平台类别', '产品系列', '产品型号'],
        columns='主推分类',
        values='总面积',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    # 确保三列都存在
    for cat in ['标准化', '配置化', '定制化']:
        if cat not in pivot_df.columns:
            pivot_df[cat] = 0

    pivot_df['面积汇总'] = pivot_df[['标准化', '配置化', '定制化']].sum(axis=1)

    # 合并定制原因（从grouped中获取，每个型号唯一）
    custom_map = grouped[['平台类别', '产品系列', '产品型号', '定制原因']].drop_duplicates()
    pivot_df = pivot_df.merge(custom_map, on=['平台类别', '产品系列', '产品型号'], how='left')

    # 调整列顺序：定制原因放在最后
    final_df = pivot_df[['平台类别', '产品系列', '产品型号', '面积汇总', '标准化', '配置化', '定制化', '定制原因']]
    final_df = final_df.sort_values(['平台类别', '产品系列', '产品型号']).reset_index(drop=True)

    # 4. 计算总计（用于占比）
    total_area = final_df['面积汇总'].sum()
    total_std = final_df['标准化'].sum()
    total_cfg = final_df['配置化'].sum()
    total_cus = final_df['定制化'].sum()

    # ---- 构建列标题（含占比） ----
    std_ratio = (total_std / total_area * 100) if total_area else 0
    cfg_ratio = (total_cfg / total_area * 100) if total_area else 0
    cus_ratio = (total_cus / total_area * 100) if total_area else 0
    col_std = f"标准化({std_ratio:.1f}%)"
    col_cfg = f"配置化({cfg_ratio:.1f}%)"
    col_cus = f"定制化({cus_ratio:.1f}%)"

    # 5. 构建最终DataFrame（含小计和总计）
    rows = []
    for plat, group in final_df.groupby('平台类别', sort=False):
        # 明细行
        for _, row in group.iterrows():
            rows.append(row.to_dict())
        # 小计行
        g_sum_area = group['面积汇总'].sum()
        g_sum_std = group['标准化'].sum()
        g_sum_cfg = group['配置化'].sum()
        g_sum_cus = group['定制化'].sum()
        ratio = (g_sum_area / total_area * 100) if total_area != 0 else 0
        ratio_text = f"占比 {ratio:.2f}%"
        subtotal = {
            '平台类别': f"{plat}汇总",
            '产品系列': ratio_text,
            '产品型号': '',
            '面积汇总': g_sum_area,
            '标准化': g_sum_std,
            '配置化': g_sum_cfg,
            '定制化': g_sum_cus,
            '定制原因': ''          # 小计行定制为空
        }
        rows.append(subtotal)

    # 总计行
    total_row = {
        '平台类别': '总计',
        '产品系列': '',
        '产品型号': '',
        '面积汇总': total_area,
        '标准化': total_std,
        '配置化': total_cfg,
        '定制化': total_cus,
        '定制原因': ''
    }
    rows.append(total_row)

    result_df = pd.DataFrame(rows)
    # 重命名列（包含占比的标题）
    result_df.columns = ['平台类别', '产品系列', '产品型号', '面积汇总', col_std, col_cfg, col_cus, '定制原因']

    return result_df


# ---------- 应用Excel合并单元格（9列，定制原因在最右侧） ----------
def apply_merges(sheet, start_row=2):
    """
    合并单元格并设置所有内容居中：
    列结构：平台类别(1), 产品系列(2), 产品型号(3), 面积汇总(4), 标准化(5), 配置化(6), 定制化(7), 定制原因(8)
    注意：Excel列索引从1开始
    1) 平台类别（明细行，列1）
    2) 产品系列（同一平台内连续相同，列2）
    3) 产品型号（同一产品系列内连续相同，列3）
    4) 小计行产品系列+产品型号合并（列2-3合并）
    5) 全部单元格水平垂直居中，小计/总计加粗
    """
    max_row = sheet.max_row
    if max_row < start_row:
        return

    # 设置所有单元格水平垂直居中（包括表头），共8列
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # ---------- 第一步：先收集所有明细行的原始信息（在合并之前） ----------
    detail_rows = []  # (row_idx, plat_val, series_val, model_val)
    for r in range(start_row, max_row + 1):
        plat_val = sheet.cell(row=r, column=1).value   # 平台类别在第1列
        series_val = sheet.cell(row=r, column=2).value
        model_val = sheet.cell(row=r, column=3).value
        if plat_val is None:
            continue
        if isinstance(plat_val, str) and (plat_val.endswith('汇总') or plat_val == '总计'):
            continue
        detail_rows.append((r, plat_val, series_val, model_val))

    # ---------- 第二步：加粗小计/总计，合并小计行产品系列+产品型号（列2-3） ----------
    for row_idx in range(start_row, max_row + 1):
        plat_cell = sheet.cell(row=row_idx, column=1)   # 平台类别在第1列
        if plat_cell.value and isinstance(plat_cell.value, str):
            val = plat_cell.value
            if val.endswith('汇总'):
                # 合并产品系列和产品型号（列2和3）
                sheet.merge_cells(start_row=row_idx, start_column=2,
                                  end_row=row_idx, end_column=3)
                b_val = sheet.cell(row=row_idx, column=2).value
                merged_cell = sheet.cell(row=row_idx, column=2)
                merged_cell.value = b_val
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                # 加粗该行所有单元格（8列）
                for col in range(1, 9):
                    sheet.cell(row=row_idx, column=col).font = Font(bold=True)
            elif val == '总计':
                for col in range(1, 9):
                    sheet.cell(row=row_idx, column=col).font = Font(bold=True)

    # ---------- 第三步：平台类别合并（列1，基于已收集的明细行） ----------
    if detail_rows:
        from collections import defaultdict
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))

        for plat, rows in platform_groups.items():
            start_r = rows[0][0]
            end_r = rows[-1][0]
            if len(rows) > 1:
                sheet.merge_cells(start_row=start_r, start_column=1,
                                  end_row=end_r, end_column=1)
                merged_cell = sheet.cell(row=start_r, column=1)
                merged_cell.value = plat
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ---------- 第四步：产品系列和产品型号合并（列2和列3，基于收集的明细行） ----------
    if detail_rows:
        from collections import defaultdict
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))

        for plat, rows in platform_groups.items():
            idx = 0
            while idx < len(rows):
                row_idx, series_val, model_val = rows[idx]
                if series_val is None:
                    idx += 1
                    continue
                # 找连续相同的产品系列（列2）
                end_idx = idx
                while end_idx < len(rows) and str(rows[end_idx][1]) == str(series_val):
                    end_idx += 1
                if end_idx - idx > 1:
                    start_r = rows[idx][0]
                    end_r = rows[end_idx - 1][0]
                    sheet.merge_cells(start_row=start_r, start_column=2,
                                      end_row=end_r, end_column=2)
                    merged_cell = sheet.cell(row=start_r, column=2)
                    merged_cell.value = series_val
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 在该产品系列组内，合并连续相同的产品型号（列3）
                sub_rows = rows[idx:end_idx]
                j = 0
                while j < len(sub_rows):
                    r_j, _, model_val_j = sub_rows[j]
                    if model_val_j is None:
                        j += 1
                        continue
                    j2 = j
                    while j2 < len(sub_rows) and sub_rows[j2][2] == model_val_j:
                        j2 += 1
                    if j2 - j > 1:
                        start_r_model = sub_rows[j][0]
                        end_r_model = sub_rows[j2 - 1][0]
                        sheet.merge_cells(start_row=start_r_model, start_column=3,
                                          end_row=end_r_model, end_column=3)
                        merged_cell = sheet.cell(row=start_r_model, column=3)
                        merged_cell.value = model_val_j
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    j = j2
                idx = end_idx

    # 设置列宽（8列）
    sheet.column_dimensions['A'].width = 22   # 平台类别
    sheet.column_dimensions['B'].width = 18   # 产品系列
    sheet.column_dimensions['C'].width = 12   # 产品型号
    sheet.column_dimensions['D'].width = 14   # 面积汇总
    sheet.column_dimensions['E'].width = 16   # 标准化
    sheet.column_dimensions['F'].width = 16   # 配置化
    sheet.column_dimensions['G'].width = 16   # 定制化
    sheet.column_dimensions['H'].width = 14   # 定制原因


# ---------- 生成带合并和定制原因的HTML表格（定制原因在最右侧） ----------
def df_to_html_with_merges(df, title):
    """
    将含小计/总计的 DataFrame 转换为 HTML 表格字符串，
    自动合并连续相同的 平台类别、产品系列、产品型号（仅明细行），
    小计/总计行为浅蓝色，值为0显示空。
    列顺序：平台类别, 产品系列, 产品型号, 面积汇总, 标准化, 配置化, 定制化, 定制原因
    """
    rows = df.to_dict('records')
    if not rows:
        return "<p>无数据</p>"

    cols = df.columns.tolist()
    std_col = next((c for c in cols if '标准化' in c), None)
    cfg_col = next((c for c in cols if '配置化' in c), None)
    cus_col = next((c for c in cols if '定制化' in c), None)
    area_col = '面积汇总'
    if std_col is None or cfg_col is None or cus_col is None:
        raise ValueError("找不到标准化/配置化/定制化列")

    thead = "<thead><tr>" + "".join(f"<th>{col}</th>" for col in cols) + "</tr></thead>"

    tbody_rows = []
    row_types = []
    for row in rows:
        plat = row['平台类别']
        if plat == '总计':
            row_types.append('total')
        elif isinstance(plat, str) and plat.endswith('汇总'):
            row_types.append('subtotal')
        else:
            row_types.append('detail')

    i = 0
    total_rows = len(rows)

    while i < total_rows:
        row = rows[i]
        row_type = row_types[i]

        if row_type in ('subtotal', 'total'):
            plat_val = row['平台类别']
            if row_type == 'subtotal':
                series_val = row['产品系列']
            else:
                series_val = ''
            custom_val = ''
            area_val = row[area_col]
            std_val = row[std_col]
            cfg_val = row[cfg_col]
            cus_val = row[cus_col]

            area_disp = '' if area_val == 0 else f"{area_val:.4f}".rstrip('0').rstrip('.')
            std_disp = '' if std_val == 0 else f"{std_val:.4f}".rstrip('0').rstrip('.')
            cfg_disp = '' if cfg_val == 0 else f"{cfg_val:.4f}".rstrip('0').rstrip('.')
            cus_disp = '' if cus_val == 0 else f"{cus_val:.4f}".rstrip('0').rstrip('.')

            css_class = 'subtotal' if row_type == 'subtotal' else 'total'
            tr = f'<tr class="{css_class}">'
            tr += f'<td>{plat_val}</td>'
            tr += f'<td colspan="2">{series_val}</td>'
            tr += f'<td>{area_disp}</td>'
            tr += f'<td>{std_disp}</td>'
            tr += f'<td>{cfg_disp}</td>'
            tr += f'<td>{cus_disp}</td>'
            tr += f'<td>{custom_val}</td>'
            tr += '</tr>'
            tbody_rows.append(tr)
            i += 1
            continue

        # 明细行
        curr_plat = row['平台类别']
        start_i = i
        while i < total_rows and row_types[i] == 'detail' and rows[i]['平台类别'] == curr_plat:
            i += 1
        end_i = i - 1
        plat_rowspan = end_i - start_i + 1

        j = start_i
        while j <= end_i:
            curr_series = rows[j]['产品系列']
            series_start = j
            while j <= end_i and rows[j]['产品系列'] == curr_series:
                j += 1
            series_end = j - 1
            series_rowspan = series_end - series_start + 1

            k = series_start
            while k <= series_end:
                curr_model = rows[k]['产品型号']
                model_start = k
                while k <= series_end and rows[k]['产品型号'] == curr_model:
                    k += 1
                model_end = k - 1
                model_rowspan = model_end - model_start + 1

                for idx in range(model_start, model_end + 1):
                    row_data = rows[idx]
                    if idx == model_start:
                        custom_val = row_data.get('定制原因', '')
                        plat_td = f'<td rowspan="{plat_rowspan}">{curr_plat}</td>' if idx == start_i else ''
                        series_td = f'<td rowspan="{series_rowspan}">{curr_series}</td>' if idx == series_start else ''
                        model_td = f'<td rowspan="{model_rowspan}">{curr_model}</td>' if idx == model_start else ''
                    else:
                        custom_val = ''
                        plat_td = ''
                        series_td = ''
                        model_td = ''

                    area_val = row_data[area_col]
                    std_val = row_data[std_col]
                    cfg_val = row_data[cfg_col]
                    cus_val = row_data[cus_col]
                    area_disp = '' if area_val == 0 else f"{area_val:.4f}".rstrip('0').rstrip('.')
                    std_disp = '' if std_val == 0 else f"{std_val:.4f}".rstrip('0').rstrip('.')
                    cfg_disp = '' if cfg_val == 0 else f"{cfg_val:.4f}".rstrip('0').rstrip('.')
                    cus_disp = '' if cus_val == 0 else f"{cus_val:.4f}".rstrip('0').rstrip('.')

                    tr = '<tr>'
                    if plat_td:
                        tr += plat_td
                    if series_td:
                        tr += series_td
                    if model_td:
                        tr += model_td
                    tr += f'<td>{area_disp}</td>'
                    tr += f'<td>{std_disp}</td>'
                    tr += f'<td>{cfg_disp}</td>'
                    tr += f'<td>{cus_disp}</td>'
                    tr += f'<td>{custom_val}</td>'
                    tr += '</tr>'
                    tbody_rows.append(tr)

                k = model_end + 1
            j = series_end + 1
        i = end_i + 1

    html = f"""
    <div class="table-container">
        <h2>{title}</h2>
        <table>
            {thead}
            <tbody>
                {''.join(tbody_rows)}
            </tbody>
        </table>
    </div>
    """
    return html


# ---------- 主函数 ----------
def main():
    data_path = 'data/接单明细V1.1.xlsx'
    data = pd.read_excel(data_path)
    print("原始列：", data.columns.tolist())

    titles = ['创建日期', '区域', '销售大区', '省份/国家', '国家', '订单数量', '总面积',
              '订单推送金额-RMB', '签单金额（万元）', '产品系列', '产品型号', '产品间距',
              '产品主推类型（产品维度）', '修正后的产品线', '修正后的业务产品线', '品牌',
              '业务产品线', '旗舰类别', '平台类别', '定位']
    data = data[titles]

    data['主推分类'] = data['产品主推类型（产品维度）'].apply(classify_main_type)

    domestic = data[data['区域'] == '国内'].copy()
    international = data[data['区域'] == '国际'].copy()

    df_dom = process_region_data(domestic)
    df_int = process_region_data(international)

    # ---------- 1. 写入Excel（已适配8列，定制原因在最后） ----------
    output_file = '分组统计结果.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_dom.to_excel(writer, sheet_name='国内', index=False)
        df_int.to_excel(writer, sheet_name='国际', index=False)
        workbook = writer.book
        for sheet_name in ['国内', '国际']:
            apply_merges(workbook[sheet_name])

    print(f"Excel保存完成！文件：{output_file}")

    # ---------- 2. 生成网页HTML ----------
    dom_html = df_to_html_with_merges(df_dom, "国内")
    int_html = df_to_html_with_merges(df_int, "国际")

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>分组统计结果</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h2 {{ text-align: center; color: #333; }}
            .table-container {{ margin-bottom: 40px; }}
            table {{ border-collapse: collapse; width: 100%; margin: 0 auto; }}
            th, td {{ border: 1px solid #000; padding: 6px 12px; text-align: center; }}
            th {{ background-color: #f2f2f2; font-weight: bold; color: #4A90D9; }}
            .subtotal td, .total td {{ color: #4A90D9; font-weight: bold; background-color: #f9f9ff; }}
        </style>
    </head>
    <body>
        {dom_html}
        {int_html}
    </body>
    </html>
    """

    html_file = '分组统计结果.html'
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    webbrowser.open(html_file)
    print(f"网页已生成并打开：{html_file}")


if __name__ == '__main__':
    main()