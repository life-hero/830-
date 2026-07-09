# -*- coding: utf-8 -*-

"""
分组统计工具 – 图形界面版本（基于 tkinter）
支持从原始数据生成统计报表（Excel + HTML），或从已生成的 Excel 重建 HTML。
"""

import webbrowser
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
from typing import List, Dict, Any, Optional, Tuple
import json
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import time

# ---------- 常量 ----------
FONT_SIZE_HTML = 12
DEBUG = True

def debug_print(msg: str) -> None:
    if DEBUG:
        timestamp = time.strftime("%H:%M:%S")
        print(f"[{timestamp}] {msg}")

# ---------- 分类映射 ----------
def classify_main_type(val: str) -> str:
    val_lower = val.lower()
    if val_lower in ['pci-top', 'pci-top+'] or val_lower in ['top+', 'top']:
        return '标准化'
    elif val_lower == 'pci-non-top' or val_lower == 'non-top':
        return '配置化'
    elif '定制' == val:
        return '定制化'
    else:
        return '其他'

# ---------- 按销售大区统计 ----------
def compute_region_stats(df: pd.DataFrame) -> Tuple[List[str], Dict[str, Any]]:
    debug_print("开始 compute_region_stats")
    df_temp = df.copy()
    df_temp['主推分类'] = df_temp['产品主推类型（产品维度）'].apply(classify_main_type)

    region_class = df_temp.groupby(['销售大区', '主推分类'])['总面积'].sum().unstack(fill_value=0)
    for cat in ['标准化', '配置化', '定制化']:
        if cat not in region_class.columns:
            region_class[cat] = 0
    region_class['面积汇总'] = region_class.sum(axis=1)

    total_area = region_class['面积汇总'].sum()
    region_class['占比'] = (region_class['面积汇总'] / total_area * 100) if total_area else 0
    region_class_sorted = region_class.sort_values('面积汇总', ascending=False)
    region_class_sorted['排名'] = range(1, len(region_class_sorted) + 1)

    sorted_regions = region_class_sorted.index.tolist()
    region_data = {}
    for region in sorted_regions:
        row = region_class_sorted.loc[region]
        region_data[region] = {
            '面积汇总': row['面积汇总'],
            '标准化': row['标准化'],
            '配置化': row['配置化'],
            '定制化': row['定制化'],
            '占比': row['占比'],
            '排名': row['排名']
        }
    debug_print("compute_region_stats 完成")
    return sorted_regions, region_data


# =============================================================================
# 【新增函数】处理产品型号级别的订单详情数据
# =============================================================================
def process_order_detail_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理订单级别的明细数据，用于生成订单号、毛利情况、低毛利原因等列。

    按 平台类别+产品系列+产品型号 分组，汇总订单号列表、毛利平均值、低毛利原因明细。

    Args:
        df: 原始数据DataFrame，需包含列：
            - 平台类别
            - 产品系列
            - 产品型号
            - 客户采购订单编号
            - 毛利情况
            - 低毛利原因

    Returns:
        DataFrame: 包含列：
            - 平台类别
            - 产品系列
            - 产品型号
            - 订单号列表 (逗号分隔)
            - 毛利平均值
            - 低毛利原因明细 (按行展开)
    """

    def convert_margin(val):
        if pd.api.types.is_number(val):
            return float(val)
        s = str(val).strip()
        if any('\u4e00' <= ch <= '\u9fff' for ch in s):
            return float('nan')  # 改为 NaN，而不是 0
        if s.endswith('%'):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return float('nan')
        try:
            return float(s)
        except ValueError:
            return float('nan')

    debug_print("开始 process_order_detail_data")

    required_cols = ['平台类别', '产品系列', '产品型号', '订单号列表', '毛利平均值', '订单号',
                     '毛利值', '低毛利原因']

    if df.empty:
        return pd.DataFrame(columns=required_cols)

    source_cols = ['平台类别', '产品系列', '产品型号', '客户采购订单编号', '毛利情况', '低毛利原因']
    for col in source_cols:
        if col not in df.columns:
            debug_print(f"警告: 列 '{col}' 不存在，将创建空列")
            df[col] = ''

    df_temp = df.copy()
    df_temp['客户采购订单编号'] = df_temp['客户采购订单编号'].fillna('').astype(str)
    df_temp['低毛利原因'] = df_temp['低毛利原因'].fillna('').astype(str)
    df_temp['毛利情况'] = df_temp['毛利情况'].apply(convert_margin)
    # 不填充 NaN，保留以便后续处理

    grouped = df_temp.groupby(['平台类别', '产品系列', '产品型号'], as_index=False)

    def agg_order_numbers(group):
        orders = group['客户采购订单编号'].unique()
        orders = [o for o in orders if o and o != '']
        return ', '.join(orders) if orders else ''

    def agg_profit_avg(group):
        # 【修复】使用 pd.to_numeric 安全转换
        profits = pd.to_numeric(group['毛利情况'], errors='coerce')
        return profits.mean() if not profits.isna().all() else 0.0

    def agg_low_profit_reasons(group):
        result = []
        for order, sub_group in group.groupby('客户采购订单编号'):
            if order and order != '':
                # 【修复】安全转换
                profit_values = pd.to_numeric(sub_group['毛利情况'], errors='coerce')
                profit_avg = profit_values.mean() if not profit_values.isna().all() else 0.0
                reasons = sub_group['低毛利原因'].unique()
                reasons = [r for r in reasons if r and r != '']
                if reasons:
                    for reason in reasons:
                        result.append({
                            '订单号': order,
                            '毛利值': profit_avg,
                            '低毛利原因': reason
                        })
                else:
                    result.append({
                        '订单号': order,
                        '毛利值': profit_avg,
                        '低毛利原因': ''
                    })
        return result

    result_rows = []
    for (plat, series, model), group in grouped:
        orders = agg_order_numbers(group)
        profit_avg = agg_profit_avg(group)
        reasons_detail = agg_low_profit_reasons(group)

        if reasons_detail:
            for item in reasons_detail:
                result_rows.append({
                    '平台类别': plat,
                    '产品系列': series,
                    '产品型号': model,
                    '订单号列表': orders,
                    '毛利平均值': profit_avg,
                    '订单号': item['订单号'],
                    '毛利值': item['毛利值'],
                    '低毛利原因': item['低毛利原因']
                })
        else:
            result_rows.append({
                '平台类别': plat,
                '产品系列': series,
                '产品型号': model,
                '订单号列表': orders,
                '毛利平均值': profit_avg,
                '订单号': '',
                '毛利值': '',
                '低毛利原因': ''
            })

    result_df = pd.DataFrame(result_rows)

    for col in required_cols:
        if col not in result_df.columns:
            result_df[col] = ''
            debug_print(f"添加缺失列: {col}")

    debug_print(f"process_order_detail_data 完成，行数：{len(result_df)}")
    return result_df


# =============================================================================
# 【修改函数】核心数据处理 - 增加订单明细维度
# =============================================================================
def process_region_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理区域数据，生成包含订单明细的统计报表。
    """
    debug_print("开始 process_region_data")

    # =========================================================================
    # 按 平台类别+产品系列+产品型号+订单号 分组汇总面积
    # =========================================================================
    if '客户采购订单编号' not in df.columns:
        df['客户采购订单编号'] = ''
        debug_print("警告: 列 '客户采购订单编号' 不存在，将创建空列")

    df_temp = df.copy()
    df_temp['客户采购订单编号'] = df_temp['客户采购订单编号'].fillna('').astype(str)

    grouped = df_temp.groupby(
        ['平台类别', '产品系列', '产品型号', '客户采购订单编号', '主推分类'],
        as_index=False
    ).agg({
        '总面积': 'sum'
    })

    pivot_df = grouped.pivot_table(
        index=['平台类别', '产品系列', '产品型号', '客户采购订单编号'],
        columns='主推分类',
        values='总面积',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    for cat in ['标准化', '配置化', '定制化']:
        if cat not in pivot_df.columns:
            pivot_df[cat] = 0

    pivot_df['面积汇总'] = pivot_df[['标准化', '配置化', '定制化']].sum(axis=1)
    pivot_df['订单号'] = pivot_df['客户采购订单编号']

    # ---------- 毛利情况处理 ----------
    if '毛利情况' in df_temp.columns:
        # 【修复】先转换毛利情况为数值
        df_temp['毛利情况_数值'] = pd.to_numeric(df_temp['毛利情况'], errors='coerce')

        profit_group = df_temp.groupby(
            ['平台类别', '产品系列', '产品型号', '客户采购订单编号'],
            as_index=False
        )['毛利情况_数值'].mean()
        profit_group.rename(columns={'毛利情况_数值': '毛利平均值'}, inplace=True)
        pivot_df = pivot_df.merge(profit_group,
                                  on=['平台类别', '产品系列', '产品型号', '客户采购订单编号'],
                                  how='left')
        pivot_df['毛利平均值'] = pivot_df['毛利平均值'].fillna(0).round(2)
        pivot_df['毛利平均值'] = (pivot_df['毛利平均值'] * 100).round(2)
    else:
        pivot_df['毛利平均值'] = 0

    # ---------- 低毛利原因处理 ----------
    if '低毛利原因' in df_temp.columns:
        reason_group = df_temp.groupby(
            ['平台类别', '产品系列', '产品型号', '客户采购订单编号'],
            as_index=False
        )['低毛利原因'].first()
        pivot_df = pivot_df.merge(reason_group,
                                  on=['平台类别', '产品系列', '产品型号', '客户采购订单编号'],
                                  how='left')
        pivot_df['低毛利原因'] = pivot_df['低毛利原因'].fillna('')
    else:
        pivot_df['低毛利原因'] = ''


    # 配置化理由（空列）
    pivot_df['配置化理由'] = ''

    # 定制原因处理
    custom_df = df_temp[df_temp['产品主推类型（产品维度）'] == '定制'][
        ['产品型号']].drop_duplicates()
    custom_df['定制原因'] = '定制'
    pivot_df = pivot_df.merge(custom_df, on='产品型号', how='left')
    pivot_df['定制原因'] = pivot_df['定制原因'].fillna('')

    # ---------- 区域统计 ----------
    region_stats = compute_region_stats(df)
    sorted_regions, region_data = region_stats

    region_group = df_temp.groupby(
        ['平台类别', '产品系列', '产品型号', '客户采购订单编号', '销售大区', '主推分类'],
        as_index=False
    ).agg({
        '总面积': 'sum'
    })

    region_pivot = region_group.pivot_table(
        index=['平台类别', '产品系列', '产品型号', '客户采购订单编号'],
        columns=['销售大区', '主推分类'],
        values='总面积',
        fill_value=0
    ).reset_index()

    key_cols = ['平台类别', '产品系列', '产品型号', '客户采购订单编号']
    keys = pd.DataFrame()
    for col in key_cols:
        found = False
        for c in region_pivot.columns:
            if col in str(c):
                keys[col] = region_pivot[c]
                found = True
                break
        if not found:
            raise KeyError(f"在 region_pivot 中找不到列: {col}")

    stat_cols = region_pivot.drop(columns=list(keys.columns))
    stat_cols.columns = [
        f"{col[0]}_{col[1]}" if isinstance(col, tuple) else str(col)
        for col in stat_cols.columns
    ]
    region_pivot_flat = pd.concat([keys, stat_cols], axis=1)

    pivot_df = pivot_df.merge(region_pivot_flat, on=key_cols, how='left')

    region_cols = []
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            region_cols.append(f"{region}_{cat}")

    for col in region_cols:
        if col not in pivot_df.columns:
            pivot_df[col] = 0
        else:
            pivot_df[col] = pivot_df[col].fillna(0)

    pivot_df['总计'] = pivot_df[region_cols].sum(axis=1)

    # ---------- 数值列处理 ----------
    numeric_base_cols = ['面积汇总', '标准化', '配置化', '定制化'] + region_cols + ['总计']
    for col in numeric_base_cols:
        if col in pivot_df.columns:
            pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce').fillna(0)
            pivot_df[col] = pivot_df[col].round(0).astype(int)

    pivot_df['面积汇总'] = pivot_df['标准化'] + pivot_df['配置化'] + pivot_df['定制化']

    if '客户采购订单编号' in pivot_df.columns:
        pivot_df = pivot_df.drop(columns=['客户采购订单编号'])
        debug_print("删除多余的客户采购订单编号列")

    # ---------- 排序 ----------
    pivot_df = pivot_df.sort_values(
        ['平台类别', '产品系列', '产品型号', '订单号'],
        ascending=[True, True, True, True]
    ).reset_index(drop=True)

    # ---------- 构建最终结果 ----------
    total_area = pivot_df['面积汇总'].sum()
    total_std = pivot_df['标准化'].sum()
    total_cfg = pivot_df['配置化'].sum()
    total_cus = pivot_df['定制化'].sum()

    std_ratio = (total_std / total_area * 100) if total_area else 0
    cfg_ratio = (total_cfg / total_area * 100) if total_area else 0
    cus_ratio = (total_cus / total_area * 100) if total_area else 0

    col_std = f"标准化({std_ratio:.1f}%)"
    col_cfg = f"配置化({cfg_ratio:.1f}%)"
    col_cus = f"定制化({cus_ratio:.1f}%)"

    rows = []

    # 总计行
    total_row_dict = {
        '平台类别': '总计',
        '产品系列': '',
        '产品型号': '',
        '订单号': '',
        '毛利情况': '',
        '低毛利原因': '',
        '面积汇总': total_area,
        '标准化': total_std,
        '配置化': total_cfg,
        '配置化理由': '',
        '定制化': total_cus,
        '定制原因': '',
    }
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            col_name = f"{region}_{cat}"
            total_row_dict[col_name] = pivot_df[col_name].sum()
    total_row_dict['总计'] = pivot_df['总计'].sum()

    # 顶部总计行
    top_total_row = total_row_dict.copy()
    top_total_row['平台类别'] = ''
    top_total_row['产品系列'] = ''
    top_total_row['产品型号'] = ''
    rows.append(top_total_row)

    # 按平台分组处理
    for plat, group in pivot_df.groupby('平台类别', sort=False):
        # 按产品系列分组
        for series, series_group in group.groupby('产品系列', sort=False):
            # 按产品型号分组 - 只输出明细行，不生成型号汇总
            for model, model_group in series_group.groupby('产品型号', sort=False):
                # 每个订单号单独一行
                for _, row in model_group.iterrows():
                    row_dict = row.to_dict()
                    # 确保所有列都存在
                    for col in ['订单号', '毛利平均值', '配置化理由', '低毛利原因']:
                        if col not in row_dict:
                            row_dict[col] = ''
                    # 重命名毛利平均值 -> 毛利情况
                    row_dict['毛利情况'] = row_dict.get('毛利平均值', '')
                    if '毛利平均值' in row_dict:
                        del row_dict['毛利平均值']
                    rows.append(row_dict)

                # 【已删除】产品型号汇总行

            # 【已删除】产品系列汇总行 - 不再生成

        # 平台汇总行
        p_sum_area = group['面积汇总'].sum()
        p_sum_std = group['标准化'].sum()
        p_sum_cfg = group['配置化'].sum()
        p_sum_cus = group['定制化'].sum()
        ratio = (p_sum_area / total_area * 100) if total_area != 0 else 0
        ratio_text = f"占比 {ratio:.2f}%"

        platform_subtotal = {
            '平台类别': f"{plat}汇总",
            '产品系列': ratio_text,
            '产品型号': '',
            '订单号': '',
            '毛利情况': '',
            '低毛利原因': '',
            '面积汇总': p_sum_area,
            '标准化': p_sum_std,
            '配置化': p_sum_cfg,
            '配置化理由': '',
            '定制化': p_sum_cus,
            '定制原因': '',
        }
        for region in sorted_regions:
            for cat in ['标准化', '配置化', '定制化']:
                col_name = f"{region}_{cat}"
                if col_name in group.columns:
                    platform_subtotal[col_name] = group[col_name].sum()
                else:
                    platform_subtotal[col_name] = 0
        platform_subtotal['总计'] = sum(platform_subtotal.get(col, 0) for col in region_cols)
        rows.append(platform_subtotal)

    # 底部总计行
    rows.append(total_row_dict)

    result_df = pd.DataFrame(rows)

    # ---------- 列重命名 ----------
    base_col_names = [
        '平台类别', '产品系列', '产品型号',
        '订单号', '毛利情况(%)', '低毛利原因',
        '面积汇总', col_std, col_cfg, '配置化理由', col_cus, '定制原因'
    ]

    region_col_names = []
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            region_col_names.append(f"{region}_{cat}")

    new_cols = base_col_names + region_col_names + ['总计']

    # 确保列数量匹配
    if len(new_cols) != len(result_df.columns):
        debug_print(
            f"警告: 列数量不匹配, new_cols={len(new_cols)}, df_cols={len(result_df.columns)}")
        while len(new_cols) < len(result_df.columns):
            new_cols.append(f"extra_col_{len(new_cols)}")
        new_cols = new_cols[:len(result_df.columns)]

    result_df.columns = new_cols

    # 存储属性
    result_df.attrs['sorted_regions'] = sorted_regions
    result_df.attrs['region_data'] = region_data
    result_df.attrs['region_area'] = {r: region_data[r]['面积汇总'] for r in sorted_regions}
    result_df.attrs['region_ratio'] = {r: region_data[r]['占比'] for r in sorted_regions}
    result_df.attrs['region_rank'] = {r: region_data[r]['排名'] for r in sorted_regions}

    debug_print("process_region_data 完成")
    return result_df

# =============================================================================
# 【修改函数】Excel合并与样式 - 适配新增列
# =============================================================================
def apply_merges_bk(
    sheet,
    start_row: int = 3,
    sorted_regions: List[str] = None,
    region_data: Dict[str, Any] = None,
    col_names: List[str] = None
) -> None:
    """
    应用Excel合并和样式。

    适配新增列：
        - 订单号列（索引8）
        - 毛利情况列（索引9）
        - 配置化理由列（索引10）
        - 低毛利订单号（索引11）
        - 低毛利值（索引12）
        - 低毛利原因（索引13）

    Args:
        sheet: openpyxl工作表对象
        start_row: 数据起始行
        sorted_regions: 排序后的区域列表
        region_data: 区域数据字典
        col_names: 列名列表
    """
    debug_print("开始 apply_merges")
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row < start_row:
        return

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 基础列名（前14列包含新增列）
    if col_names is None:
        col_names = [sheet.cell(row=1, column=i).value for i in range(1, 15)]
        if not any(col_names):
            col_names = [
                '平台类别', '产品系列', '产品型号', '面积汇总',
                '标准化', '配置化', '定制化', '定制原因',
                '订单号', '毛利情况', '配置化理由',
                '低毛利订单号', '低毛利值', '低毛利原因'
            ]

    # ---------- 1. 设置表头（前14列合并两行） ----------
    # 基础列：前14列
    for col_idx, name in enumerate(col_names[:14], start=1):
        sheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = name
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # ---------- 2. 区域列（原有逻辑） ----------
    if sorted_regions and region_data:
        start_col = 15  # 前14列为固定列，第15列开始为区域列
        total_col_index = start_col + len(sorted_regions) * 3

        for i, region in enumerate(sorted_regions):
            col_start = start_col + i * 3
            col_end = col_start + 2
            sheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            info = region_data.get(region, {})
            rank = info.get('排名', 0)
            area = info.get('面积汇总', 0)
            ratio = info.get('占比', 0)
            display_text = f"TOP{rank:.0f} {region} - {area:.2f}㎡ ({ratio:.1f}%)"
            cell = sheet.cell(row=1, column=col_start)
            cell.value = display_text
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            for j, sub in enumerate(['标准化', '配置化', '定制化']):
                cell = sheet.cell(row=2, column=col_start + j)
                cell.value = sub
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # 总计列
        sheet.merge_cells(start_row=1, start_column=total_col_index, end_row=2, end_column=total_col_index)
        cell = sheet.cell(row=1, column=total_col_index)
        cell.value = '总计'
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # ---------- 3. 设置所有单元格边框和对齐 ----------
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # ---------- 4. 总计行样式（第3行） ----------
    if max_row >= 3:
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
        cell = sheet.cell(row=3, column=1)
        cell.value = '总计'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(4, max_col + 1):
            cell = sheet.cell(row=3, column=col)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)

    # ---------- 5. 收集明细行数据 ----------
    detail_rows = []
    for r in range(start_row, max_row + 1):
        plat_val = sheet.cell(row=r, column=1).value
        if plat_val is None:
            continue
        if isinstance(plat_val, str) and (plat_val.endswith('汇总') or plat_val == '总计'):
            continue
        if r == 3:
            continue
        series_val = sheet.cell(row=r, column=2).value
        model_val = sheet.cell(row=r, column=3).value
        detail_rows.append((r, plat_val, series_val, model_val))

    # ---------- 6. 汇总行和总计行样式 ----------
    for row_idx in range(start_row, max_row + 1):
        plat_cell = sheet.cell(row=row_idx, column=1)
        if plat_cell.value and isinstance(plat_cell.value, str):
            val = plat_cell.value
            if val.endswith('汇总'):
                # 先获取值再合并
                b_val = sheet.cell(row=row_idx, column=2).value
                # 合并单元格（合并后只保留左上角单元格）
                sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
                # 合并后，只有左上角单元格(column=2)可以赋值
                merged_cell = sheet.cell(row=row_idx, column=2)
                merged_cell.value = b_val
                # 设置样式时，跳过已被合并的列（3列已被合并到2列）
                for col in range(1, max_col + 1):
                    try:
                        cell = sheet.cell(row=row_idx, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA",
                                                fill_type="solid")
                    except AttributeError:
                        # 跳过 MergedCell 的样式设置（MergedCell 没有 font 和 fill 属性）
                        continue
            elif val == '总计':
                # 先获取值再合并
                sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                merged_cell = sheet.cell(row=row_idx, column=1)
                merged_cell.value = '总计'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                for col in range(1, max_col + 1):
                    try:
                        cell = sheet.cell(row=row_idx, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                                fill_type="solid")
                    except AttributeError:
                        # 跳过 MergedCell 的样式设置
                        continue

    # ---------- 7. 平台类别合并 ----------
    if detail_rows:
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))
        for plat, rows in platform_groups.items():
            start_r = rows[0][0]
            end_r = rows[-1][0]
            if len(rows) > 1:
                sheet.merge_cells(start_row=start_r, start_column=1,
                                  end_row=end_r, end_column=1)
                merged_cell = sheet.cell(row=start_r, column=1)
                merged_cell.value = plat

    # ---------- 8. 产品系列合并 ----------
    if detail_rows:
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))
        for plat, rows in platform_groups.items():
            idx = 0
            while idx < len(rows):
                row_idx, series_val, model_val = rows[idx]
                if series_val is None:
                    idx += 1
                    continue
                end_idx = idx
                while end_idx < len(rows) and str(rows[end_idx][1]) == str(series_val):
                    end_idx += 1
                if end_idx - idx > 1:
                    start_r = rows[idx][0]
                    end_r = rows[end_idx - 1][0]
                    sheet.merge_cells(start_row=start_r, start_column=2,
                                      end_row=end_r, end_column=2)
                    merged_cell = sheet.cell(row=start_r, column=2)
                    merged_cell.value = series_val
                sub_rows = rows[idx:end_idx]
                j = 0
                while j < len(sub_rows):
                    r_j, _, model_val_j = sub_rows[j]
                    if model_val_j is None:
                        j += 1
                        continue
                    j2 = j
                    while j2 < len(sub_rows) and sub_rows[j2][2] == model_val_j:
                        j2 += 1
                    if j2 - j > 1:
                        start_r_model = sub_rows[j][0]
                        end_r_model = sub_rows[j2 - 1][0]
                        sheet.merge_cells(start_row=start_r_model, start_column=3,
                                          end_row=end_r_model, end_column=3)
                        merged_cell = sheet.cell(row=start_r_model, column=3)
                        merged_cell.value = model_val_j
                    j = j2
                idx = end_idx

    # ---------- 9. 设置列宽 ----------
    col_widths = {
        1: 22,   # 平台类别
        2: 18,   # 产品系列
        3: 14,   # 产品型号
        4: 14,   # 面积汇总
        5: 16,   # 标准化
        6: 16,   # 配置化
        7: 16,   # 定制化
        8: 14,   # 定制原因
        9: 18,   # 订单号
        10: 14,  # 毛利情况
        11: 16,  # 配置化理由
        12: 18,  # 低毛利订单号
        13: 14,  # 低毛利值
        14: 20   # 低毛利原因
    }
    for col, width in col_widths.items():
        if col <= max_col:
            sheet.column_dimensions[get_column_letter(col)].width = width

    # 区域列宽
    for col in range(15, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14

    debug_print("apply_merges 完成")


def _safe_set_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    """
    安全设置单元格样式，跳过 MergedCell。

    Args:
        cell: openpyxl 单元格对象
        font: Font 对象
        fill: PatternFill 对象
        alignment: Alignment 对象
        border: Border 对象

    Returns:
        bool: 是否成功设置样式
    """
    try:
        # 检查是否为 MergedCell（通过尝试访问 value 属性）
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        return True
    except AttributeError:
        # MergedCell 没有这些属性，跳过
        return False
    except TypeError:
        # 某些单元格可能不支持某些属性
        return False



def apply_merges(
    sheet,
    start_row: int = 3,
    sorted_regions: List[str] = None,
    region_data: Dict[str, Any] = None,
    col_names: List[str] = None
) -> None:
    """应用Excel合并和样式 - 适配12列基础结构"""
    debug_print("开始 apply_merges")
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row < start_row:
        return

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 【修复】基础列名：12列（与process_region_data输出一致）
    if col_names is None:
        col_names = [sheet.cell(row=1, column=i).value for i in range(1, 13)]
        if not any(col_names) or col_names[0] is None:
            col_names = [
                '平台类别', '产品系列', '产品型号',
                '订单号', '毛利情况(%)', '低毛利原因',
                '面积汇总', '标准化', '配置化', '配置化理由', '定制化', '定制原因'
            ]

    # ---------- 1. 设置表头（前12列合并两行） ----------
    for col_idx, name in enumerate(col_names[:12], start=1):
        sheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = name
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # ---------- 2. 区域列（第13列开始） ----------
    if sorted_regions and region_data:
        start_col = 13  # 12列基础列 + 1
        total_col_index = start_col + len(sorted_regions) * 3

        for i, region in enumerate(sorted_regions):
            col_start = start_col + i * 3
            col_end = col_start + 2
            sheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
            info = region_data.get(region, {})
            rank = info.get('排名', 0)
            area = info.get('面积汇总', 0)
            ratio = info.get('占比', 0)
            display_text = f"TOP{rank:.0f} {region} - {area:.2f}㎡ ({ratio:.1f}%)"
            cell = sheet.cell(row=1, column=col_start)
            cell.value = display_text
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            for j, sub in enumerate(['标准化', '配置化', '定制化']):
                cell = sheet.cell(row=2, column=col_start + j)
                cell.value = sub
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # 总计列
        sheet.merge_cells(start_row=1, start_column=total_col_index, end_row=2, end_column=total_col_index)
        cell = sheet.cell(row=1, column=total_col_index)
        cell.value = '总计'
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # ---------- 3. 设置所有单元格边框和对齐 ----------
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # ---------- 4. 总计行样式（第3行） ----------
    if max_row >= 3:
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
        cell = sheet.cell(row=3, column=1)
        cell.value = '总计'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(4, max_col + 1):
            cell = sheet.cell(row=3, column=col)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)

    # ---------- 5. 收集明细行数据 ----------
    detail_rows = []
    for r in range(start_row, max_row + 1):
        plat_val = sheet.cell(row=r, column=1).value
        if plat_val is None:
            continue
        if isinstance(plat_val, str) and (plat_val.endswith('汇总') or plat_val == '总计'):
            continue
        if r == 3:
            continue
        series_val = sheet.cell(row=r, column=2).value
        model_val = sheet.cell(row=r, column=3).value
        detail_rows.append((r, plat_val, series_val, model_val))

    # ---------- 6. 汇总行和总计行样式 ----------
    for row_idx in range(start_row, max_row + 1):
        plat_cell = sheet.cell(row=row_idx, column=1)
        if plat_cell.value and isinstance(plat_cell.value, str):
            val = plat_cell.value
            if val.endswith('汇总'):
                # 获取系列值
                b_val = sheet.cell(row=row_idx, column=2).value
                sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
                merged_cell = sheet.cell(row=row_idx, column=2)
                merged_cell.value = b_val
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif val == '总计':
                sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                merged_cell = sheet.cell(row=row_idx, column=1)
                merged_cell.value = '总计'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ---------- 7. 平台类别合并 ----------
    if detail_rows:
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))
        for plat, rows in platform_groups.items():
            start_r = rows[0][0]
            end_r = rows[-1][0]
            if len(rows) > 1:
                sheet.merge_cells(start_row=start_r, start_column=1,
                                  end_row=end_r, end_column=1)
                merged_cell = sheet.cell(row=start_r, column=1)
                merged_cell.value = plat

    # ---------- 8. 产品系列合并 ----------
    if detail_rows:
        platform_groups = defaultdict(list)
        for r, plat, series, model in detail_rows:
            platform_groups[plat].append((r, series, model))
        for plat, rows in platform_groups.items():
            idx = 0
            while idx < len(rows):
                row_idx, series_val, model_val = rows[idx]
                if series_val is None:
                    idx += 1
                    continue
                end_idx = idx
                while end_idx < len(rows) and str(rows[end_idx][1]) == str(series_val):
                    end_idx += 1
                if end_idx - idx > 1:
                    start_r = rows[idx][0]
                    end_r = rows[end_idx - 1][0]
                    sheet.merge_cells(start_row=start_r, start_column=2,
                                      end_row=end_r, end_column=2)
                    merged_cell = sheet.cell(row=start_r, column=2)
                    merged_cell.value = series_val
                sub_rows = rows[idx:end_idx]
                j = 0
                while j < len(sub_rows):
                    r_j, _, model_val_j = sub_rows[j]
                    if model_val_j is None:
                        j += 1
                        continue
                    j2 = j
                    while j2 < len(sub_rows) and sub_rows[j2][2] == model_val_j:
                        j2 += 1
                    if j2 - j > 1:
                        start_r_model = sub_rows[j][0]
                        end_r_model = sub_rows[j2 - 1][0]
                        sheet.merge_cells(start_row=start_r_model, start_column=3,
                                          end_row=end_r_model, end_column=3)
                        merged_cell = sheet.cell(row=start_r_model, column=3)
                        merged_cell.value = model_val_j
                    j = j2
                idx = end_idx

    # ---------- 9. 设置列宽 ----------
    # 12列基础列宽
    col_widths = {
        1: 22,   # 平台类别
        2: 18,   # 产品系列
        3: 14,   # 产品型号
        4: 18,   # 订单号
        5: 14,   # 毛利情况(%)
        6: 20,   # 低毛利原因
        7: 14,   # 面积汇总
        8: 16,   # 标准化
        9: 16,   # 配置化
        10: 16,  # 配置化理由
        11: 16,  # 定制化
        12: 14   # 定制原因
    }
    for col, width in col_widths.items():
        if col <= max_col:
            sheet.column_dimensions[get_column_letter(col)].width = width

    # 区域列宽
    for col in range(13, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14

    debug_print("apply_merges 完成")



# =============================================================================
# HTML生成 - 适配新增列
# =============================================================================

def df_to_html_with_merges(df: pd.DataFrame, title: str, font_size: int = FONT_SIZE_HTML) -> str:
    """
    将DataFrame转换为HTML表格，保持与Excel一致的列顺序和数据。
    """
    debug_print(f"开始 df_to_html_with_merges for {title}")

    df = df.copy()
    # 处理文本列的空值
    for col in ['平台类别', '产品系列', '产品型号', '定制原因', '订单号', '配置化理由',
                '低毛利订单号', '低毛利原因']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str)

    # 【修复】毛利情况和低毛利值保留为数值类型，不转字符串
    for col in ['毛利情况(%)', '低毛利值']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    rows = df.to_dict('records')
    if not rows:
        return "<p>无数据</p>"

    cols = df.columns.tolist()
    sorted_regions = df.attrs.get('sorted_regions', [])
    region_data = df.attrs.get('region_data', {})

    # ---------- 识别列 ----------
    order_col = '订单号'
    profit_col = '毛利情况(%)'
    low_reason_col = '低毛利原因'
    area_col = '面积汇总'
    custom_col = '定制原因'
    reason_col = '配置化理由'
    total_col = '总计'

    # 三大分类列（通过列名查找）
    std_col = None
    cfg_col = None
    cus_col = None
    for c in cols:
        if c not in ['平台类别', '产品系列', '产品型号', '订单号', '毛利情况(%)', '低毛利原因',
                     '面积汇总', '配置化理由', '定制原因', '总计']:
            if not any(c.startswith(r) for r in sorted_regions):
                if '标准化' in c:
                    std_col = c
                elif '配置化' in c:
                    cfg_col = c
                elif '定制化' in c:
                    cus_col = c

    # ---------- 辅助函数：格式化单元格值 ----------
    def format_value(val):
        """格式化单元格值，0或空值显示为空"""
        if val is None or val == '':
            return ''
        if isinstance(val, (int, float)):
            if val == 0:
                return ''
            return str(int(val))
        return str(val)

    def format_percent(val):
        """格式化百分比值"""
        if val is None or val == '':
            return ''
        try:
            num_val = float(val)
            if pd.isna(num_val) or num_val == 0:
                return ''
            return f'{num_val:.2f}%'
        except (ValueError, TypeError):
            return str(val)

    # ---------- 构建表头 ----------
    base_headers = [
        '平台类别', '产品系列', '产品型号',
        order_col, profit_col, low_reason_col,
        area_col, std_col, cfg_col, reason_col, cus_col, custom_col
    ]

    thead_parts = ['<thead><tr>']
    for col_name in base_headers:
        thead_parts.append(f'<th rowspan="2">{col_name}</th>')

    for region in sorted_regions:
        info = region_data.get(region, {'排名': 0, '面积汇总': 0, '占比': 0})
        rank = info['排名']
        area = info['面积汇总']
        ratio = info['占比']
        display_text = f"TOP{rank:.0f} - {region}<br>{area:.2f}㎡ - ({ratio:.1f}%)"
        thead_parts.append(f'<th colspan="3">{display_text}</th>')

    thead_parts.append(f'<th rowspan="2">{total_col}</th>')
    thead_parts.append('</tr><tr>')

    for _ in sorted_regions:
        thead_parts.append('<th>标准化</th><th>配置化</th><th>定制化</th>')

    thead_parts.append('</tr></thead>')
    thead = ''.join(thead_parts)

    # ---------- 构建表体 ----------
    tbody_rows = []
    row_types = []
    for row in rows:
        plat = row.get('平台类别', '')
        if plat == '总计':
            row_types.append('total')
        elif isinstance(plat, str) and plat.endswith('汇总'):
            row_types.append('subtotal')
        elif plat == '':
            row_types.append('top_total')
        else:
            row_types.append('detail')

    i = 0
    total_rows = len(rows)

    def get_row_value(row, col_name):
        """安全获取行数据"""
        if col_name is None:
            return ''
        return row.get(col_name, '')

    while i < total_rows:
        row = rows[i]
        row_type = row_types[i]

        if row_type == 'top_total':
            tr = '<tr class="top-total">'
            tr += '<td colspan="3">总计</td>'
            # 订单号
            tr += f'<td>{format_value(get_row_value(row, order_col))}</td>'
            # 毛利情况（百分比）
            tr += f'<td>{format_percent(get_row_value(row, profit_col))}</td>'
            # 低毛利原因
            tr += f'<td>{format_value(get_row_value(row, low_reason_col))}</td>'
            # 面积汇总
            tr += f'<td>{format_value(get_row_value(row, area_col))}</td>'
            # 标准化
            tr += f'<td>{format_value(get_row_value(row, std_col))}</td>'
            # 配置化
            tr += f'<td>{format_value(get_row_value(row, cfg_col))}</td>'
            # 配置化理由
            tr += f'<td>{format_value(get_row_value(row, reason_col))}</td>'
            # 定制化
            tr += f'<td>{format_value(get_row_value(row, cus_col))}</td>'
            # 定制原因
            tr += f'<td>{format_value(get_row_value(row, custom_col))}</td>'
            # 区域列
            for region in sorted_regions:
                for cat in ['标准化', '配置化', '定制化']:
                    val = row.get(f"{region}_{cat}", 0)
                    tr += f'<td>{format_value(val)}</td>'
            total_val = row.get(total_col, 0)
            tr += f'<td>{format_value(total_val)}</td>'
            tr += '</tr>'
            tbody_rows.append(tr)
            i += 1
            continue

        if row_type in ('subtotal', 'total'):
            plat_val = row.get('平台类别', '')
            series_val = row.get('产品系列', '')
            css_class = 'subtotal' if row_type == 'subtotal' else 'total'

            tr = f'<tr class="{css_class}">'
            if row_type == 'total':
                tr += f'<td colspan="3">{plat_val}</td>'
            else:
                tr += f'<td>{plat_val}</td>'
                tr += f'<td colspan="2">{series_val}</td>'

            # 订单号
            tr += f'<td>{format_value(get_row_value(row, order_col))}</td>'
            # 毛利情况（百分比）
            tr += f'<td>{format_percent(get_row_value(row, profit_col))}</td>'
            # 低毛利原因
            tr += f'<td>{format_value(get_row_value(row, low_reason_col))}</td>'
            # 面积汇总
            tr += f'<td>{format_value(get_row_value(row, area_col))}</td>'
            # 标准化
            tr += f'<td>{format_value(get_row_value(row, std_col))}</td>'
            # 配置化
            tr += f'<td>{format_value(get_row_value(row, cfg_col))}</td>'
            # 配置化理由
            tr += f'<td>{format_value(get_row_value(row, reason_col))}</td>'
            # 定制化
            tr += f'<td>{format_value(get_row_value(row, cus_col))}</td>'
            # 定制原因
            tr += f'<td>{format_value(get_row_value(row, custom_col))}</td>'

            # 区域列
            for region in sorted_regions:
                for cat in ['标准化', '配置化', '定制化']:
                    val = row.get(f"{region}_{cat}", 0)
                    tr += f'<td>{format_value(val)}</td>'
            total_val = row.get(total_col, 0)
            tr += f'<td>{format_value(total_val)}</td>'
            tr += '</tr>'
            tbody_rows.append(tr)
            i += 1
            continue

        # ---------- 明细行 ----------
        curr_plat = row.get('平台类别', '')
        start_i = i
        while i < total_rows and row_types[i] == 'detail' and rows[i].get('平台类别',
                                                                          '') == curr_plat:
            i += 1
        end_i = i - 1
        plat_rowspan = end_i - start_i + 1

        j = start_i
        while j <= end_i:
            curr_series = rows[j].get('产品系列', '')
            series_start = j
            while j <= end_i and rows[j].get('产品系列', '') == curr_series:
                j += 1
            series_end = j - 1
            series_rowspan = series_end - series_start + 1

            k = series_start
            while k <= series_end:
                curr_model = rows[k].get('产品型号', '')
                model_start = k
                while k <= series_end and rows[k].get('产品型号', '') == curr_model:
                    k += 1
                model_end = k - 1
                model_rowspan = model_end - model_start + 1

                for idx in range(model_start, model_end + 1):
                    row_data = rows[idx]
                    tr = '<tr>'

                    if idx == start_i:
                        tr += f'<td rowspan="{plat_rowspan}">{curr_plat}</td>'
                    if idx == series_start:
                        tr += f'<td rowspan="{series_rowspan}">{curr_series}</td>'
                    if idx == model_start:
                        tr += f'<td rowspan="{model_rowspan}">{curr_model}</td>'

                    # 订单号
                    tr += f'<td>{format_value(get_row_value(row_data, order_col))}</td>'
                    # 毛利情况（百分比）
                    tr += f'<td>{format_percent(get_row_value(row_data, profit_col))}</td>'
                    # 低毛利原因
                    tr += f'<td>{format_value(get_row_value(row_data, low_reason_col))}</td>'
                    # 面积汇总
                    tr += f'<td>{format_value(get_row_value(row_data, area_col))}</td>'
                    # 标准化
                    tr += f'<td>{format_value(get_row_value(row_data, std_col))}</td>'
                    # 配置化
                    tr += f'<td>{format_value(get_row_value(row_data, cfg_col))}</td>'
                    # 配置化理由
                    tr += f'<td>{format_value(get_row_value(row_data, reason_col))}</td>'
                    # 定制化
                    tr += f'<td>{format_value(get_row_value(row_data, cus_col))}</td>'
                    # 定制原因
                    tr += f'<td>{format_value(get_row_value(row_data, custom_col))}</td>'

                    # 区域列
                    for region in sorted_regions:
                        for cat in ['标准化', '配置化', '定制化']:
                            val = row_data.get(f"{region}_{cat}", 0)
                            tr += f'<td>{format_value(val)}</td>'
                    total_val = row_data.get(total_col, 0)
                    tr += f'<td>{format_value(total_val)}</td>'
                    tr += '</tr>'
                    tbody_rows.append(tr)

                k = model_end + 1
            j = series_end + 1
        i = end_i + 1

    # ---------- 生成HTML ----------
    html = f"""
    <div class="table-container">
        <h2>{title}</h2>
        <table style="font-size: {font_size}px; border-collapse: collapse; width: 100%;">
            {thead}
            <tbody>
                {''.join(tbody_rows)}
            </tbody>
        </table>
    </div>
    """
    debug_print(f"df_to_html_with_merges for {title} 完成")
    return html


# =============================================================================
# 从Excel重建DataFrame - 适配新增列
# =============================================================================


def rebuild_df_from_excel(excel_path: str, sheet_name: str) -> pd.DataFrame:
    """
    从Excel重建DataFrame。

    适配新增列：订单号、毛利情况(%)、低毛利原因、配置化理由

    Args:
        excel_path: Excel文件路径
        sheet_name: Sheet名称（'国内' 或 '国际'）

    Returns:
        DataFrame
    """
    debug_print(f"开始 rebuild_df_from_excel: {excel_path}, sheet={sheet_name}")

    meta_df = pd.read_excel(excel_path, sheet_name='元数据', header=None, index_col=0)
    if sheet_name == '国内':
        regions_key = 'sorted_regions_dom'
        data_key = 'region_data_dom'
    elif sheet_name == '国际':
        regions_key = 'sorted_regions_int'
        data_key = 'region_data_int'
    else:
        raise ValueError(f"不支持的 sheet_name: {sheet_name}，仅支持 '国内' 或 '国际'")

    try:
        sorted_regions = json.loads(meta_df.loc[regions_key, 1])
        region_data = json.loads(meta_df.loc[data_key, 1])
    except KeyError as e:
        raise KeyError(f"元数据中缺少键: {e}，请确认 Excel 是由本工具生成。")

    # 读取数据，跳过前2行（表头）
    df_data = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, skiprows=2)

    # 列数 = 12基础列 + 3*区域数 + 1总计
    num_cols = 12 + 3 * len(sorted_regions) + 1
    df_data = df_data.iloc[:, :num_cols]

    # 读取表头（第一行）
    header_row = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=1).iloc[0]

    # 【修改】基础列名（前12列）：适配新列顺序
    base_cols_from_excel = [str(header_row[i]) for i in range(12)]

    # 区域列
    region_cols = []
    for region in sorted_regions:
        for cat in ['标准化', '配置化', '定制化']:
            region_cols.append(f"{region}_{cat}")

    final_cols = base_cols_from_excel + region_cols + ['总计']
    df_data.columns = final_cols

    # 填充空值（平台类别、产品系列、产品型号）
    for col in ['平台类别', '产品系列', '产品型号']:
        if col in df_data.columns:
            df_data[col] = df_data[col].ffill()

    # 数值列处理（排除文本列）
    text_cols = ['平台类别', '产品系列', '产品型号', '订单号', '毛利情况(%)',
                 '低毛利原因', '配置化理由', '定制原因']
    numeric_cols = [c for c in final_cols if c not in text_cols]
    for col in numeric_cols:
        df_data[col] = pd.to_numeric(df_data[col], errors='coerce').fillna(0)

    # 【修改】计算面积汇总：从第6列（索引6）获取
    # 新列顺序：0平台类别,1产品系列,2产品型号,3订单号,4毛利情况(%),5低毛利原因,6面积汇总,7标准化,8配置化,9配置化理由,10定制化,11定制原因
    area_col = base_cols_from_excel[6] if len(base_cols_from_excel) > 6 else '面积汇总'
    if area_col not in df_data.columns:
        # 备用：从标准化+配置化+定制化计算
        std_col = base_cols_from_excel[7] if len(base_cols_from_excel) > 7 else None
        cfg_col = base_cols_from_excel[8] if len(base_cols_from_excel) > 8 else None
        cus_col = base_cols_from_excel[10] if len(base_cols_from_excel) > 10 else None
        if std_col and cfg_col and cus_col:
            df_data['面积汇总'] = df_data[std_col] + df_data[cfg_col] + df_data[cus_col]
        else:
            df_data['面积汇总'] = 0

    # 存储属性
    df_data.attrs['sorted_regions'] = sorted_regions
    df_data.attrs['region_data'] = region_data
    df_data.attrs['region_area'] = {r: region_data[r]['面积汇总'] for r in sorted_regions}
    df_data.attrs['region_ratio'] = {r: region_data[r]['占比'] for r in sorted_regions}
    df_data.attrs['region_rank'] = {r: region_data[r]['排名'] for r in sorted_regions}

    debug_print(f"rebuild_df_from_excel 完成，行数：{len(df_data)}")
    return df_data



# =============================================================================
# 平台汇总数据提取 - 适配新增列
# =============================================================================
def _extract_platform_data(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    从单个DataFrame中提取平台级别的三大分类面积汇总。

    Args:
        df: 由 process_region_data 生成的DataFrame

    Returns:
        DataFrame包含列：平台类别, 标准化面积, 配置化面积, 定制化面积
        若没有明细数据则返回None
    """
    if df is None or df.empty:
        return None

    debug_print("_extract_platform_data: 开始提取")

    # 【修复】使用固定索引获取三大分类列
    # 列顺序: 平台类别(0), 产品系列(1), 产品型号(2), 订单号(3), 毛利情况(4),
    #         低毛利原因(5), 面积汇总(6), 标准化(7), 配置化(8), 配置化理由(9),
    #         定制化(10), 定制原因(11), 区域列(12+), 总计(最后)
    try:
        std_col = df.columns[7]   # 标准化列
        cfg_col = df.columns[8]   # 配置化列
        cus_col = df.columns[10]  # 定制化列
    except IndexError as e:
        debug_print(f"_extract_platform_data: 获取列索引失败: {e}")
        debug_print(f"可用列: {df.columns.tolist()}")
        return None

    debug_print(f"_extract_platform_data: std_col={std_col}, cfg_col={cfg_col}, cus_col={cus_col}")

    # 过滤明细行（排除汇总行和总计行）
    mask = (df['平台类别'].notna()) & (~df['平台类别'].astype(str).str.endswith('汇总')) & (df['平台类别'] != '总计') & (df['平台类别'] != '')
    df_detail = df.loc[mask, ['平台类别', std_col, cfg_col, cus_col]].copy()

    if df_detail.empty:
        debug_print("_extract_platform_data: 无明细数据")
        return None

    # 转为数值
    df_detail[std_col] = pd.to_numeric(df_detail[std_col], errors='coerce').fillna(0)
    df_detail[cfg_col] = pd.to_numeric(df_detail[cfg_col], errors='coerce').fillna(0)
    df_detail[cus_col] = pd.to_numeric(df_detail[cus_col], errors='coerce').fillna(0)

    # 按平台汇总
    grouped = df_detail.groupby('平台类别', as_index=False).agg({
        std_col: 'sum',
        cfg_col: 'sum',
        cus_col: 'sum'
    })
    grouped.rename(columns={std_col: '标准化面积', cfg_col: '配置化面积', cus_col: '定制化面积'}, inplace=True)

    debug_print(f"_extract_platform_data: 完成, 平台数={len(grouped)}")
    return grouped

# =============================================================================
# GUI应用程序类（保持不变，仅更新内部方法）
# =============================================================================
class GUIApp:
    """分组统计工具图形界面主类"""

    def __init__(self, master: tk.Tk) -> None:
        self.master = master
        master.title("分组统计工具")
        master.geometry("800x800")
        master.resizable(True, True)

        self.data_file_path = tk.StringVar()
        self.excel_file_path = tk.StringVar()
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        self.btn_stat = None
        self.btn_convert = None
        self.btn_add_data = None
        self.btn_select_excel = None

        self._converted_df_dom = None
        self._converted_df_int = None
        self._converted_excel_path = None

        self._platform_summary_df = None
        self._region_summary_df = None
        self._converted_platform_df = None
        self._converted_region_df = None

        self._create_widgets()

    def _create_widgets(self) -> None:
        top_frame = tk.Frame(self.master)
        top_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        hint_label = tk.Label(top_frame, text="使用说明-Copyright 2026 Cuiyj", font=('Arial', 12, 'bold'))
        hint_label.pack(anchor='w')

        hint_text = scrolledtext.ScrolledText(
            top_frame,
            wrap=tk.WORD,
            font=('Arial', 13),
            height=10,
            relief=tk.SUNKEN,
            borderwidth=2
        )
        hint_text.pack(fill=tk.BOTH, expand=True)

        info = (
            "【数据文件要求】\n"
            "请选择包含原始数据的 Excel 文件（.xlsx 或 .xls），该文件必须包含以下列：\n"
            "  - 区域         : 取值为 '国内' 或 '国际'，用于区分国内/国际统计。\n"
            "  - 销售大区     : 如华东、华南等，用于按大区统计。\n"
            "  - 省份/国家    : 省份或国家名称（仅用于展示）。\n"
            "  - 国家         : 国家名称（仅用于展示）。\n"
            "  - 总面积       : 数值型，面积数据。\n"
            "  - 产品系列     : 产品所属系列。\n"
            "  - 产品型号     : 具体型号。\n"
            "  - 产品主推类型（产品维度）: 取值为 'TOP'、'TOP+'、'NON-TOP' 或 '定制'。\n"
            "  - 平台类别     : 如 '320×160平台'、'户内专显600平台' 等。\n"
            "  - 客户采购订单编号 : 订单编号。\n"
            "  - 毛利情况     : 毛利数据。\n"
            "  - 低毛利原因   : 低毛利原因描述。\n"
            "Excel 兼容格式：.xlsx 或 .xls。\n\n"
            "【操作说明】\n"
            " 第一步：先生成Excel格式的统计数据 \n"
            "   1.1. 点击 '添加数据文件' 按钮选择原始数据 Excel。\n"
            "   1.2. 点击 '统计' 按钮生成分组统计 Excel 和 HTML 文件，并弹出保存对话框。\n\n"
            " 第二步：选中修改后的Excel文件，将其转成html文件 \n"
            "   2.1. 对于已修改好的Excel（仅限本工具生成的 Excel 格式）文件，可使用下方的 '选择 Excel 文件' 按钮加载。\n"
            "   2.2. 点击 'Excel 转 HTML' 按钮将选中的 Excel 文件转换为 HTML。\n\n"
            " ***重要提醒：统计生成的Excel文件内容可以修改，格式不能调整，否则无法转成Html文件。*** \n"
        )
        hint_text.insert(tk.END, info)
        hint_text.config(state=tk.DISABLED)

        bottom_frame = tk.Frame(self.master)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)

        group1 = tk.LabelFrame(bottom_frame, text="第一步：从原始数据统计出分类数据", padx=5, pady=5)
        group1.pack(fill=tk.X, pady=5)

        row1 = tk.Frame(group1)
        row1.pack(fill=tk.X, pady=2)

        tk.Label(row1, text="数据文件:", width=10, anchor='w').pack(side=tk.LEFT)
        entry_data = tk.Entry(row1, textvariable=self.data_file_path, width=35, state='readonly')
        entry_data.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        self.btn_add_data = tk.Button(row1, text="添加接单明细数据文件", command=self._select_data_file)
        self.btn_add_data.pack(side=tk.LEFT, padx=2)

        self.btn_stat = tk.Button(row1, text="生成统计数据", command=self._run_statistics)
        self.btn_stat.pack(side=tk.LEFT, padx=2)

        group2 = tk.LabelFrame(bottom_frame, text="第二步：将修改后的Excel 转 HTML", padx=5, pady=5)
        group2.pack(fill=tk.X, pady=5)

        row2 = tk.Frame(group2)
        row2.pack(fill=tk.X, pady=2)

        tk.Label(row2, text="Excel 文件:", width=10, anchor='w').pack(side=tk.LEFT)
        entry_excel = tk.Entry(row2, textvariable=self.excel_file_path, width=40, state='readonly')
        entry_excel.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        btn_frame = tk.Frame(row2)
        btn_frame.pack(side=tk.RIGHT, padx=5)

        self.btn_select_excel = tk.Button(btn_frame, text="选择 Excel 文件", width=10,
                                          command=self._select_excel_file)
        self.btn_select_excel.pack(side=tk.LEFT, padx=2)

        self.btn_convert = tk.Button(btn_frame, text="Excel 转 HTML", width=15,
                                     command=self._convert_excel_to_html)
        self.btn_convert.pack(side=tk.LEFT, padx=2, ipadx=10)

        status_bar = tk.Label(self.master, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def _select_data_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="选择原始数据文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.data_file_path.set(file_path)

    def _select_excel_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="选择分组统计 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.excel_file_path.set(file_path)

    def _run_statistics(self) -> None:
        data_path = self.data_file_path.get()
        if not data_path:
            messagebox.showerror("错误", "请先选择数据文件。")
            return

        self._set_buttons_state(False)
        self.status_var.set("正在处理统计任务...")
        debug_print("统计任务开始")

        def task():
            try:
                self._do_statistics(data_path)
            except Exception as e:
                error_msg = str(e)
                self.master.after(0, lambda: self._show_error(f"统计失败：{error_msg}"))
            finally:
                self.master.after(0, self._finish_task)

        threading.Thread(target=task, daemon=True).start()

    def _do_statistics(self, data_path: str) -> None:
        debug_print(f"读取原始数据: {data_path}")
        data = pd.read_excel(data_path)

        required_cols = ['区域', '销售大区', '省份/国家', '国家', '总面积',
                         '产品系列', '产品型号', '产品主推类型（产品维度）', '平台类别']
        missing = [c for c in required_cols if c not in data.columns]
        if missing:
            raise ValueError(f"数据文件缺少必需列：{', '.join(missing)}")

        # 【新增】检查新增列是否存在，不存在则创建空列
        optional_cols = ['客户采购订单编号', '毛利情况', '低毛利原因']
        for col in optional_cols:
            if col not in data.columns:
                data[col] = ''
                debug_print(f"创建空列: {col}")

        data['主推分类'] = data['产品主推类型（产品维度）'].apply(classify_main_type)

        domestic = data[data['区域'] == '国内'].copy()
        international = data[data['区域'] == '国际'].copy()

        if domestic.empty and international.empty:
            raise ValueError("数据中既无 '国内' 也无 '国际' 记录。")

        df_dom = process_region_data(domestic) if not domestic.empty else None
        df_int = process_region_data(international) if not international.empty else None

        debug_print("统计完成，准备弹出保存对话框")
        self.master.after(0, lambda: self._save_dialog(df_dom, df_int))

    def _save_dialog(self, df_dom: Optional[pd.DataFrame], df_int: Optional[pd.DataFrame]) -> None:
        default_dir = get_download_folder()
        file_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            initialdir=default_dir,
            initialfile="分组统计结果.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not file_path:
            self.status_var.set("已取消保存")
            self._finish_task()
            return

        try:
            debug_print(f"保存 Excel: {file_path}")
            self._save_excel(file_path, df_dom, df_int)
            html_path = os.path.splitext(file_path)[0] + ".html"
            debug_print(f"生成 HTML: {html_path}")
            self._generate_html(html_path, df_dom, df_int)
            self.master.after(0, lambda: self._show_success(f"统计完成！\nExcel：{file_path}\nHTML：{html_path}"))
            self.master.after(0, lambda: webbrowser.open(html_path))
        except Exception as e:
            error_msg = str(e)
            self.master.after(0, lambda: self._show_error(f"保存Excel失败：{error_msg}"))
        finally:
            self.master.after(0, self._finish_task)


    def _extract_platform_data(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """提取平台汇总数据（适配新增列）"""
        return _extract_platform_data(self, df)

    def _add_platform_summary_sheet(self, writer: pd.ExcelWriter,
                                     df_dom: Optional[pd.DataFrame],
                                     df_int: Optional[pd.DataFrame]) -> None:
        """生成平台占比情况Sheet"""
        dom_plat = self._extract_platform_data(df_dom)
        int_plat = self._extract_platform_data(df_int)

        if dom_plat is not None and int_plat is not None:
            merged = pd.concat([dom_plat, int_plat], ignore_index=True)
            platform_df = merged.groupby('平台类别', as_index=False).sum()
        elif dom_plat is not None:
            platform_df = dom_plat.copy()
        elif int_plat is not None:
            platform_df = int_plat.copy()
        else:
            return

        platform_df['标品面积'] = platform_df['标准化面积'] + platform_df['配置化面积']
        platform_df['总计'] = platform_df['标准化面积'] + platform_df['配置化面积'] + platform_df['定制化面积']

        platform_df['标准化占比'] = (platform_df['标准化面积'] / platform_df['总计'] * 100).round(2)
        platform_df['配置占比'] = (platform_df['配置化面积'] / platform_df['总计'] * 100).round(2)
        platform_df['标品占比'] = (platform_df['标品面积'] / platform_df['总计'] * 100).round(2)
        platform_df['定制化占比'] = (platform_df['定制化面积'] / platform_df['总计'] * 100).round(2)

        total_all = platform_df['总计'].sum()
        platform_df['平台占比'] = (platform_df['总计'] / total_all * 100).round(2) if total_all else 0

        platform_df = platform_df.sort_values('总计', ascending=False).reset_index(drop=True)

        total_row = {
            '平台类别': '总计',
            '标准化面积': platform_df['标准化面积'].sum(),
            '配置化面积': platform_df['配置化面积'].sum(),
            '定制化面积': platform_df['定制化面积'].sum(),
            '标品面积': platform_df['标品面积'].sum(),
            '总计': platform_df['总计'].sum(),
        }
        total_sum = total_row['总计']
        if total_sum > 0:
            total_row['标准化占比'] = (total_row['标准化面积'] / total_sum * 100).round(2)
            total_row['配置占比'] = (total_row['配置化面积'] / total_sum * 100).round(2)
            total_row['标品占比'] = (total_row['标品面积'] / total_sum * 100).round(2)
            total_row['定制化占比'] = (total_row['定制化面积'] / total_sum * 100).round(2)
            total_row['平台占比'] = 100.0
        else:
            total_row['标准化占比'] = 0.0
            total_row['配置占比'] = 0.0
            total_row['标品占比'] = 0.0
            total_row['定制化占比'] = 0.0
            total_row['平台占比'] = 100.0

        platform_df = pd.concat([platform_df, pd.DataFrame([total_row])], ignore_index=True)

        col_order = ['平台类别', '标准化面积', '标准化占比', '配置化面积', '配置占比',
                     '标品面积', '标品占比', '定制化面积', '定制化占比', '总计', '平台占比']
        platform_df = platform_df[col_order]

        workbook = writer.book
        sheet_name = '平台占比情况'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name, 0)

        numeric_cols = platform_df.columns[1:]
        for col in numeric_cols:
            platform_df[col] = platform_df[col].replace(0, None)

        for r_idx, row in enumerate(platform_df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        headers = col_order
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for c_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(col_order)):
            for cell in row:
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = thin_border
                header_cell = sheet.cell(row=1, column=cell.column)
                if header_cell.value and '占比' in header_cell.value:
                    cell.number_format = '0.00"%"'
                else:
                    cell.number_format = '#,##0'

        col_widths = {'平台类别': 18, '标准化面积': 14, '标准化占比': 14, '配置化面积': 14,
                      '配置占比': 14, '标品面积': 14, '标品占比': 14, '定制化面积': 14,
                      '定制化占比': 14, '总计': 14, '平台占比': 14}
        for c_idx, col_name in enumerate(col_order, start=1):
            sheet.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 14)

        last_row = sheet.max_row
        for col in range(1, len(col_order) + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        self._platform_summary_df = platform_df
        debug_print("平台占比情况 Sheet 生成完成")

    def _add_region_summary_sheet(self, writer: pd.ExcelWriter,
                                   df_dom: Optional[pd.DataFrame],
                                   df_int: Optional[pd.DataFrame]) -> None:
        """生成国内国际定制情况Sheet"""

        def extract_region_totals(df: pd.DataFrame) -> Dict[str, float]:
            if df is None or df.empty:
                return {'标准化': 0, '配置化': 0, '定制化': 0, '总计': 0}

            debug_print("extract_region_totals: 开始提取")

            # 过滤明细行
            mask = (df['平台类别'].notna()) & (~df['平台类别'].astype(str).str.endswith('汇总')) & (
                        df['平台类别'] != '总计') & (df['平台类别'] != '')
            df_detail = df.loc[mask].copy()
            if df_detail.empty:
                debug_print("extract_region_totals: 无明细数据")
                return {'标准化': 0, '配置化': 0, '定制化': 0, '总计': 0}

            # 【修复】使用固定索引获取三大分类列
            # 列顺序: 平台类别(0), 产品系列(1), 产品型号(2), 订单号(3), 毛利情况(4),
            #         低毛利原因(5), 面积汇总(6), 标准化(7), 配置化(8), 配置化理由(9),
            #         定制化(10), 定制原因(11), 区域列(12+), 总计(最后)
            try:
                std_col = df.columns[7]  # 标准化列
                cfg_col = df.columns[8]  # 配置化列
                cus_col = df.columns[10]  # 定制化列
            except IndexError as e:
                debug_print(f"extract_region_totals: 获取列索引失败: {e}")
                debug_print(f"可用列: {df.columns.tolist()}")
                return {'标准化': 0, '配置化': 0, '定制化': 0, '总计': 0}

            debug_print(
                f"extract_region_totals: std_col={std_col}, cfg_col={cfg_col}, cus_col={cus_col}")

            total_std = pd.to_numeric(df_detail[std_col], errors='coerce').sum()
            total_cfg = pd.to_numeric(df_detail[cfg_col], errors='coerce').sum()
            total_cus = pd.to_numeric(df_detail[cus_col], errors='coerce').sum()

            debug_print(
                f"extract_region_totals: total_std={total_std}, total_cfg={total_cfg}, total_cus={total_cus}")

            return {'标准化': total_std, '配置化': total_cfg, '定制化': total_cus,
                    '总计': total_std + total_cfg + total_cus}


        dom_data = extract_region_totals(df_dom)
        int_data = extract_region_totals(df_int)

        rows = []
        if dom_data['总计'] > 0:
            total_dom = dom_data['总计']
            rows.append({
                '区域': '国内',
                '标配': dom_data['标准化'],
                '标配占比': dom_data['标准化'] / total_dom * 100 if total_dom else 0,
                '配置化': dom_data['配置化'],
                '配置化占比': dom_data['配置化'] / total_dom * 100 if total_dom else 0,
                '定制化': dom_data['定制化'],
                '定制化占比': dom_data['定制化'] / total_dom * 100 if total_dom else 0,
                '总计': total_dom,
                '签单占比': 0
            })
        if int_data['总计'] > 0:
            total_int = int_data['总计']
            rows.append({
                '区域': '国际',
                '标配': int_data['标准化'],
                '标配占比': int_data['标准化'] / total_int * 100 if total_int else 0,
                '配置化': int_data['配置化'],
                '配置化占比': int_data['配置化'] / total_int * 100 if total_int else 0,
                '定制化': int_data['定制化'],
                '定制化占比': int_data['定制化'] / total_int * 100 if total_int else 0,
                '总计': total_int,
                '签单占比': 0
            })

        total_all = dom_data['总计'] + int_data['总计']
        if total_all > 0:
            rows.append({
                '区域': '总计',
                '标配': dom_data['标准化'] + int_data['标准化'],
                '标配占比': (dom_data['标准化'] + int_data['标准化']) / total_all * 100 if total_all else 0,
                '配置化': dom_data['配置化'] + int_data['配置化'],
                '配置化占比': (dom_data['配置化'] + int_data['配置化']) / total_all * 100 if total_all else 0,
                '定制化': dom_data['定制化'] + int_data['定制化'],
                '定制化占比': (dom_data['定制化'] + int_data['定制化']) / total_all * 100 if total_all else 0,
                '总计': total_all,
                '签单占比': 100.0
            })

        if not rows:
            return

        df_summary = pd.DataFrame(rows)

        total_all = df_summary[df_summary['区域'] != '总计']['总计'].sum() if not df_summary[df_summary['区域'] != '总计'].empty else 0
        if total_all > 0:
            mask_non_total = df_summary['区域'] != '总计'
            df_summary.loc[mask_non_total, '签单占比'] = (df_summary.loc[mask_non_total, '总计'] / total_all * 100).round(2)
        df_summary.loc[df_summary['区域'] == '总计', '签单占比'] = 100.0

        numeric_cols = ['标配', '配置化', '定制化', '总计']
        for col in numeric_cols:
            df_summary[col] = df_summary[col].round(0).astype(int)
        percent_cols = ['标配占比', '配置化占比', '定制化占比', '签单占比']
        for col in percent_cols:
            df_summary[col] = df_summary[col].round(2)

        workbook = writer.book
        sheet_name = '国内国际定制情况'
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        sheet = workbook.create_sheet(sheet_name, 1)

        headers = ['区域', '标配', '标配占比', '配置化', '配置化占比', '定制化', '定制化占比', '总计', '签单占比']
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for c_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        numeric_cols = df_summary.columns[1:]
        for col in numeric_cols:
            df_summary[col] = df_summary[col].replace(0, None)

        for r_idx, row in enumerate(df_summary.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = thin_border
                header_cell = sheet.cell(row=1, column=cell.column)
                if header_cell.value and '占比' in header_cell.value:
                    cell.number_format = '0.00"%"'
                else:
                    cell.number_format = '#,##0'

        col_widths = {'区域': 16, '标配': 12, '标配占比': 12, '配置化': 12,
                      '配置化占比': 12, '定制化': 12, '定制化占比': 12, '总计': 12, '签单占比': 12}
        for c_idx, col_name in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 12)

        last_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        self._region_summary_df = df_summary
        debug_print("国内国际定制情况 Sheet 生成完成")

    def _save_excel(self, file_path: str, df_dom: Optional[pd.DataFrame],
                    df_int: Optional[pd.DataFrame]) -> None:
        """保存Excel文件"""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            workbook = writer.book

            # 平台占比情况
            if df_dom is not None or df_int is not None:
                if '平台占比情况' in workbook.sheetnames:
                    workbook.remove(workbook['平台占比情况'])
                workbook.create_sheet('平台占比情况', 0)
                self._add_platform_summary_sheet(writer, df_dom, df_int)

            # 国内国际定制情况
            if df_dom is not None or df_int is not None:
                self._add_region_summary_sheet(writer, df_dom, df_int)

            # 写入国内和国际数据
            if df_dom is not None:
                df_dom_excel = df_dom.copy()
                # 所有数值列
                numeric_cols = [c for c in df_dom_excel.columns if c not in ['平台类别', '产品系列', '产品型号',
                                                                             '定制原因', '订单号', '毛利情况(%)',
                                                                             '配置化理由', '低毛利订单号',
                                                                             '低毛利值', '低毛利原因']]
                for col in numeric_cols:
                    df_dom_excel[col] = df_dom_excel[col].replace(0, None)
                df_dom_excel.to_excel(writer, sheet_name='国内', index=False, header=False, startrow=2)

            if df_int is not None:
                df_int_excel = df_int.copy()
                numeric_cols = [c for c in df_int_excel.columns if c not in ['平台类别', '产品系列', '产品型号',
                                                                             '定制原因', '订单号', '毛利情况(%)',
                                                                             '配置化理由', '低毛利订单号',
                                                                             '低毛利值', '低毛利原因']]
                for col in numeric_cols:
                    df_int_excel[col] = df_int_excel[col].replace(0, None)
                df_int_excel.to_excel(writer, sheet_name='国际', index=False, header=False, startrow=2)

            # 元数据
            meta_data = {}
            if df_dom is not None:
                meta_data['sorted_regions_dom'] = df_dom.attrs['sorted_regions']
                meta_data['region_data_dom'] = df_dom.attrs['region_data']
            if df_int is not None:
                meta_data['sorted_regions_int'] = df_int.attrs['sorted_regions']
                meta_data['region_data_int'] = df_int.attrs['region_data']

            meta_df = pd.DataFrame({
                'key': list(meta_data.keys()),
                'value': [json.dumps(v) for v in meta_data.values()]
            })
            meta_df.to_excel(writer, sheet_name='元数据', index=False, header=False)

            if '元数据' in workbook.sheetnames:
                meta_sheet = workbook['元数据']
                meta_sheet.sheet_state = 'hidden'

            # 应用样式
            for sheet_name in ['国内', '国际']:
                if sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                if sheet_name == '国内' and df_dom is not None:
                    sorted_regions = df_dom.attrs.get('sorted_regions', [])
                    region_data = df_dom.attrs.get('region_data', {})
                    col_names = df_dom.columns.tolist()
                elif sheet_name == '国际' and df_int is not None:
                    sorted_regions = df_int.attrs.get('sorted_regions', [])
                    region_data = df_int.attrs.get('region_data', {})
                    col_names = df_int.columns.tolist()
                else:
                    continue
                apply_merges(sheet, start_row=3, sorted_regions=sorted_regions,
                             region_data=region_data, col_names=col_names)

    def _df_to_simple_html(self, df: pd.DataFrame, title: str,
                           left_align_cols: List[str] = None) -> str:
        """将单行表头的DataFrame转换为HTML"""
        if df is None or df.empty:
            return ""
        if left_align_cols is None:
            left_align_cols = [df.columns[0]]

        headers = df.columns.tolist()
        thead = '<thead><tr>' + ''.join(f'<th>{h}</th>' for h in headers) + '</tr></thead>'

        tbody_rows = []
        total_indices = df[df.iloc[:, 0] == '总计'].index
        for idx, row in df.iterrows():
            tr_class = 'total' if idx in total_indices else ''
            tr = f'<tr class="{tr_class}">'
            for col in headers:
                val = row[col]
                if pd.isna(val):
                    display = ''
                elif isinstance(val, (int, float)) and val == 0:
                    display = ''
                elif '占比' in col or col.endswith('占比'):
                    display = f"{val:.2f}%" if isinstance(val, (int, float)) else str(val)
                else:
                    display = f"{int(val):,}" if isinstance(val, (int, float)) else str(val)
                align = 'left' if col in left_align_cols else 'right'
                tr += f'<td style="text-align: {align};">{display}</td>'
            tr += '</tr>'
            tbody_rows.append(tr)
        tbody = '<tbody>' + ''.join(tbody_rows) + '</tbody>'

        return f"""
        <div class="table-container">
            <h2>{title}</h2>
            <table style="font-size: {FONT_SIZE_HTML}px; border-collapse: collapse; width: 100%;">
                {thead}
                {tbody}
            </table>
        </div>
        """

    def _generate_html(self, html_path: str, df_dom: Optional[pd.DataFrame],
                       df_int: Optional[pd.DataFrame]) -> None:
        """生成HTML文件"""
        debug_print("生成HTML内容")

        platform_df = self._converted_platform_df if self._converted_platform_df is not None else self._platform_summary_df
        region_df = self._converted_region_df if self._converted_region_df is not None else self._region_summary_df

        platform_html = self._df_to_simple_html(platform_df, "平台占比情况",
                                                left_align_cols=['平台类别'])
        region_html = self._df_to_simple_html(region_df, "国内国际定制情况",
                                              left_align_cols=['区域'])
        dom_html = df_to_html_with_merges(df_dom, "国内",
                                          font_size=FONT_SIZE_HTML) if df_dom is not None else ""
        int_html = df_to_html_with_merges(df_int, "国际",
                                          font_size=FONT_SIZE_HTML) if df_int is not None else ""

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>分组统计结果</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h2 {{ text-align: center; color: #333; }}
                .table-container {{ margin-bottom: 40px; overflow-x: auto; }}
                table {{ border-collapse: collapse; width: 100%; margin: 0 auto; }}
                th, td {{ border: 1px solid #000; padding: 6px 12px; }}
                th {{ background-color: #4A90D9; color: white; font-weight: bold; text-align: center; }}
                .total td {{ background-color: #D9E1F2; font-weight: bold; }}
                .top-total td {{ background-color: #D9E1F2; font-weight: bold; }}
                .subtotal td {{ background-color: #E2EFDA; font-weight: bold; }}
            </style>
        </head>
        <body>
            {platform_html}
            {region_html}
            {dom_html}
            {int_html}
        </body>
        </html>
        """
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        debug_print(f"HTML保存完成: {html_path}")

    def _convert_excel_to_html(self) -> None:
        excel_path = self.excel_file_path.get()
        if not excel_path:
            messagebox.showerror("错误", "请先选择 Excel 文件。")
            return
        if not os.path.exists(excel_path):
            messagebox.showerror("错误", "文件不存在。")
            return
        if not excel_path.endswith('.xlsx'):
            messagebox.showerror("错误", "仅支持 .xlsx 格式。")
            return

        self._set_buttons_state(False)
        self.status_var.set("正在读取数据...")
        debug_print("Excel转HTML任务开始")

        self._converted_df_dom = None
        self._converted_df_int = None
        self._converted_excel_path = excel_path

        def read_task():
            try:
                self._do_convert_read(excel_path)
            except Exception as e:
                # 【修复】使用字符串变量捕获异常信息
                error_msg = str(e)
                self.master.after(0, lambda: self._show_error(f"读取数据失败：{error_msg}"))
                self.master.after(0, self._finish_task)

        threading.Thread(target=read_task, daemon=True).start()

    def _do_convert_read(self, excel_path: str) -> None:
        """执行Excel读取"""
        debug_print(f"开始读取Excel: {excel_path}")
        df_dom = None
        df_int = None
        platform_df = None
        region_df = None

        try:
            with pd.ExcelFile(excel_path) as xls:
                if '元数据' not in xls.sheet_names:
                    raise ValueError("该 Excel 不是由本工具生成（缺少元数据 Sheet）。")

                if '国内' in xls.sheet_names:
                    df_dom = rebuild_df_from_excel(excel_path, '国内')
                if '国际' in xls.sheet_names:
                    df_int = rebuild_df_from_excel(excel_path, '国际')

                if '平台占比情况' in xls.sheet_names:
                    platform_df = pd.read_excel(excel_path, sheet_name='平台占比情况', header=0)
                    for col in platform_df.columns[1:]:
                        platform_df[col] = pd.to_numeric(platform_df[col], errors='coerce')
                if '国内国际定制情况' in xls.sheet_names:
                    region_df = pd.read_excel(excel_path, sheet_name='国内国际定制情况', header=0)
                    for col in region_df.columns[1:]:
                        region_df[col] = pd.to_numeric(region_df[col], errors='coerce')
        except Exception as e:
            # 重新抛出异常，由上层处理
            raise

        if df_dom is None and df_int is None:
            raise ValueError("Excel 中既无 '国内' 也无 '国际' 数据。")

        self._converted_df_dom = df_dom
        self._converted_df_int = df_int
        self._converted_platform_df = platform_df
        self._converted_region_df = region_df
        debug_print("数据读取完成，准备弹出保存对话框")
        self.master.after(0, self._show_convert_save_dialog)

    def _show_convert_save_dialog(self) -> None:
        """显示转换保存对话框"""
        excel_path = self._converted_excel_path
        df_dom = self._converted_df_dom
        df_int = self._converted_df_int

        default_dir = get_download_folder()
        base_name = os.path.splitext(os.path.basename(excel_path))[0]
        html_path = filedialog.asksaveasfilename(
            title="保存 HTML 文件",
            initialdir=default_dir,
            initialfile=base_name + ".html",
            defaultextension=".html",
            filetypes=[("HTML 文件", "*.html")]
        )
        if not html_path:
            self.status_var.set("已取消保存")
            self._finish_task()
            return

        self.status_var.set("正在生成HTML...")
        debug_print(f"准备后台生成HTML: {html_path}")

        def save_task():
            try:
                self._generate_html(html_path, df_dom, df_int)
                self.master.after(0, lambda: self._show_success(f"HTML 已保存至：{html_path}"))
                self.master.after(0, lambda: webbrowser.open(html_path))
            except Exception as e:
                self.master.after(0, lambda err=e: self._show_error(f"保存HTML失败：{err}"))
            finally:
                self.master.after(0, self._finish_task)

        threading.Thread(target=save_task, daemon=True).start()

    def _set_buttons_state(self, enabled: bool) -> None:
        state = tk.NORMAL if enabled else tk.DISABLED
        if self.btn_add_data:
            self.btn_add_data.config(state=state)
        if self.btn_stat:
            self.btn_stat.config(state=state)
        if self.btn_select_excel:
            self.btn_select_excel.config(state=state)
        if self.btn_convert:
            self.btn_convert.config(state=state)

    def _finish_task(self) -> None:
        self._set_buttons_state(True)
        self.status_var.set("就绪")
        debug_print("任务结束")

    def _show_error(self, msg: str) -> None:
        messagebox.showerror("错误", msg)
        self.status_var.set("错误")

    def _show_success(self, msg: str) -> None:
        messagebox.showinfo("成功", msg)
        self.status_var.set("完成")


# =============================================================================
# 工具函数
# =============================================================================
def get_download_folder() -> str:
    """获取下载文件夹路径"""
    if os.name == 'nt':
        return os.path.join(os.path.expanduser('~'), 'Downloads')
    else:
        return os.path.join(os.path.expanduser('~'), 'Downloads')


# =============================================================================
# 主入口
# =============================================================================
def main() -> None:
    root = tk.Tk()
    app = GUIApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()